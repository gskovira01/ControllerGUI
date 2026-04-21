def send_batch_pr_program(comm, segments, axes='ABCDE', program_label='BATCHPR'):
    """
    Generate and execute a Galil batch PR program from segment data.
    Each segment should have 'converted' pulses for each axis.
    """
    if comm is None or not hasattr(comm, 'gclib'):
        raise RuntimeError('Controller communications not initialized or gclib not available.')
    if not segments:
        raise ValueError('No segments to send.')
    # Build the program string
    lines = [f'#{program_label}']
    prev_pulses = [0] * len(axes)
    import time
    # Initialize prev_pulses to current actual position for each axis (like send_datapipe_pr)
    try:
        for axis_idx in range(len(axes)):
            axis_letter = axes[axis_idx]
            resp = comm.gclib.GCommand(f"MG _RP{axis_letter}")
            val = 0
            if isinstance(resp, str):
                for line in resp.splitlines():
                    try:
                        val = int(float(line.strip()))
                        break
                    except Exception:
                        continue
            prev_pulses[axis_idx] = val
    except Exception as e:
        print(f"[WARN] Could not initialize prev_pulses from controller: {e}")
        prev_pulses = [0] * len(axes)
    for seg in segments:
        deltas = []
        for axis_idx in range(len(axes)):
            pulses = seg['converted'][axis_idx]['pulses']
            delta = pulses - prev_pulses[axis_idx]
            deltas.append(delta)
            prev_pulses[axis_idx] = pulses
        pr_cmd = 'PR ' + ','.join(str(d) for d in deltas)
        lines.append(pr_cmd)
        lines.append(f'BG {axes}')
        lines.append(f'AM {axes}')
        time.sleep(0.02)  # 20 ms delay between commands
    lines.append('EN')
    program_str = '\n'.join(lines)
    # Save the program to a file for GDK use
    try:
        with open('batch_pr.dmc', 'w') as f:
            f.write(program_str)
        print('[DEBUG] Batch PR program saved to batch_pr.dmc')
    except Exception as e:
        print(f'[WARN] Could not save batch PR program to file: {e}')
    # Download and execute the program
    comm.gclib.GProgramDownload(program_str)
    comm.gclib.GCommand(f"XQ #{program_label}")
    return program_str

"""
================================================================================
                        CONTROLLER GUI SYSTEM
================================================================================

Author: gskovira01
Last Updated: January 15, 2026
Version: 1.3.0

PURPOSE:
    Touchscreen-friendly GUI for an 8-axis controller interface with modular numeric keypad popup.
    Features tabbed interface, dynamic command mapping, modular background polling,
    indicator lights, and robust error handling. Supports mixed-mode operation with
    Galil controllers (axes A-G) and MyActuator motors (axis H).

KEY FEATURES:
    - Tabbed interface for 8 servos (A-H)
    - Mixed-mode controller support (Galil + MyActuator)
    - Modular numeric keypad popup for value entry (see numeric_keypad.py)
    - Min/max validation for all numeric fields
    - Dynamic command mapping for multiple controller types
    - Modular background polling (see ControllerPolling.py):
        * Runs as a background thread
        * Periodically polls actual position, torque, and enable/disable status for each servo
        * Communicates results to GUI via thread-safe events
    - Indicator lights and status for each servo
    - Debug log window for sent/received commands and errors
    - Zero Position button for each axis
    - Improved error handling and user feedback

REVISION HISTORY:
    Rev 1.3.0 - January 15, 2026 - Add MyActuator motor support on axis H
        - Added mixed-mode controller support (CommMode5 for MyActuator)
        - Separate comm object (comm_h) for Servo H MyActuator motor
        - Added get_comm_for_axis() and send_axis_command() routing functions
        - Updated command sending to route axis H to MyActuator, axes A-G to Galil
        - Updated controller_config.ini with CommMode5 section and AXIS_H parameters
    
    Rev 1.2.0 - December 14, 2025 - Previous updates
        - Tabbed interface and background polling improvements

HOW TO USE:
    1. CONFIGURATION:
       - Edit controller_config.ini to set up your controllers:
         * [CommMode1] section: Galil controller IP address (for axes A-G)
         * [CommMode5] section: MyActuator Waveshare IP, port, motor_id (for axis H)
         * [AXIS_A] through [AXIS_H]: Per-axis parameters (limits, scaling, descriptions)
    
    2. RUNNING THE GUI:
       - Execute: python ControllerGUI.py
       - The GUI will initialize both controllers automatically
       - Axes A-G use Galil controller (CommMode1)
       - Axis H uses MyActuator motor (CommMode5)
    
    3. USING SERVO CONTROLS:
       - Each servo tab (1-8) provides identical interface regardless of controller type
       - Enable/Disable: Activate or deactivate the servo
       - Speed/Accel/Decel: Set motion parameters (automatically converted to pulses)
       - Absolute Position: Set target position, then click "Start Motion"
       - Relative Position: Move by offset from current position
       - Jog CW/CCW: Continuous rotation at set speed
       - Zero Position: Set current position as zero reference
    
    4. MYACTUATOR-SPECIFIC NOTES (Servo H):
       - Commands are automatically translated from Galil format to CAN protocol
       - Position range: -180° to +180° (single-turn encoder with wrapping)
       - Speed limit: 500 dps recommended (hardware max ~655 dps)
       - Resolution: 0.01 degrees
       - Connection via Waveshare CAN-to-ETH converter (TCP)
    
    5. DEBUG LOG:
       - Monitor all commands and responses in the debug window
       - Useful for troubleshooting communication issues

FUNCTIONS DEFINED IN THIS FILE:
    build_servo_tab(servo_num)
        - Builds and returns the GUI layout for a single servo tab.

    handle_servo_event(event, values)
        - Handles all servo-related button events (enable, disable, jog, set values, etc.).

    get_comm_for_axis(axis_letter)
        - Returns the appropriate comm object (comm or comm_h) for the given axis.

    send_axis_command(axis_letter, cmd)
        - Routes commands to the appropriate controller based on axis.

    get_limits(axis_letter, field)
        - Returns per-axis min/max (supports speed/accel/decel/position overrides from INI).

    compute_midpoint_speed(speed, accel, decel, mid_distance)
        - Estimates achievable speed at midpoint of a move using accel/decel constraints.

    update_mid_speed_display(window, servo_num)
        - Updates the per-servo midpoint speed indicator based on current setpoints.

    save_axis_description(axis_letter, description)
        - Persists per-axis description back into controller_config.ini when edited in the GUI.

    _refresh_description_colors(window)
        - Forces description inputs to render black-on-white regardless of theme.

    (Polling logic moved to ControllerPolling.py)
        - See ControllerPolling.start_polling_thread(window, comm):
            * Starts a background thread that polls actual position, torque, and enable/disable status for each servo and updates the GUI.

================================================================================
"""

# ============================================================================
# Imports and Configuration
# ============================================================================

# ============================================================================
# Imports and Configuration
# ============================================================================
import FreeSimpleGUI as sg
import logging
import threading  # For future use if needed
import platform   # Cross-platform OS detection and adaptation
import configparser
import os
import traceback
import json
import csv
from typing import List, Dict, Any
from communications import ControllerComm
from numeric_keypad import NumericKeypad
try:
    import openpyxl  # Used for DataPipe Excel ingest
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

# Background ALL sequence control
SEQ_THREAD = None
SEQ_STOP_EVENT = None
SEQ_RUNNING = False
SEQUENCE_STATE_FILE = os.path.join(os.path.dirname(__file__), 'sequence_state.json')
MOTION_DEFAULTS_FILE = os.path.join(os.path.dirname(__file__), 'motion_defaults.json')
STARTUP_SPEED_DEFAULT = 10.0
STARTUP_ACCEL_DECEL_DEFAULT = 10.0

# Safety: Track last command type per servo to prevent unsafe Start Motion
# Values: 'abs', 'rel', 'jog', or None
LAST_MOTION_COMMAND = [None] * 8
JOG_THREADS = [None] * 8
JOG_STOP_EVENTS = [None] * 8

# Absolute safety limit: never allow more than 360 degrees rotation from zero
ABSOLUTE_SAFETY_LIMIT_DEG = 360.0
# Require consecutive out-of-limit reads before safety stop to avoid single-sample spikes.
LIMIT_TRIP_CONFIRM_SAMPLES = 3
# Tolerance added to soft limits so readback jitter near a boundary (e.g. 0°) doesn't trigger a false stop.
LIMIT_SOFT_TOLERANCE_DEG = 2.0
# Bench-test override: disable all runtime limit-stop enforcement/popups temporarily.
SAFETY_LIMIT_STOPS_ENABLED = False
# [CHANGE 2026-03-24 12:42:00 -04:00] Axis E conservative relative-step safety cap.
AXIS_E_MAX_RELATIVE_STEP_DEG = 15.0
# [CHANGE 2026-03-24 13:24:00 -04:00] Axis E jog safety: one-shot jog step cap per click.
AXIS_E_JOG_STEP_DEG = 1.0
# [CHANGE 2026-03-27 10:15:00 -04:00] Per-servo jog amount input range (degrees).
JOG_STEP_MIN_DEG = 0.0
JOG_STEP_MAX_DEG = 5.0
JOG_STEP_DEFAULT_DEG = 1.0

# ============================================================================
# Constants and Global Variables
# =====================
# Axis scaling and units dictionary
# =====================
# Example: { 'A': {'pulses': 20000, 'degrees': 360, 'scaling': 20000/360}, ... }
# Read axis parameters from controller_config.ini
import configparser
INI_PATH = os.path.join(os.path.dirname(__file__), 'controller_config.ini')
# Default descriptors per servo; editable per-tab and echoed on ALL tab
DEFAULT_SERVO_DESCRIPTIONS = {
    1: 'Primary',
    2: 'Secondary',
    3: 'Tertiary Rotation',
    4: 'Tertiary Lift',
    5: 'Address',
    6: '',
    7: '',
    8: ''
}
AXIS_UNITS = {}
axis_ini = configparser.ConfigParser()
axis_ini.read(INI_PATH)


def sync_ini_to_yaml(ini_path):
    """Push axis parameters from controller_config.ini into tim_config.yaml on iPC400.

    Called at GUI startup so the TIM service always reflects the INI master values.
    The yaml path is read from [CommMode1] tim_yaml_path in the INI.
    Fails silently with a warning if the yaml is unreachable (e.g. iPC400 offline).
    """
    try:
        import yaml
    except ImportError:
        print('[WARNING] pyyaml not installed — skipping INI→yaml sync. Run: pip install pyyaml')
        return

    ini = configparser.ConfigParser()
    ini.read(ini_path)

    yaml_path = ini.get('CommMode1', 'tim_yaml_path', fallback=None)
    if not yaml_path:
        local_yaml = os.path.join(os.path.dirname(ini_path), 'tim_service', 'tim_config.yaml')
        yaml_path = local_yaml

    try:
        with open(yaml_path, 'r') as f:
            cfg = yaml.safe_load(f) or {}
    except FileNotFoundError:
        print(f'[WARNING] tim_config.yaml not found at {yaml_path} — skipping sync')
        return
    except Exception as e:
        print(f'[WARNING] Could not read tim_config.yaml at {yaml_path}: {e}')
        return

    # Sync Axes A-D into rapidcode.axes
    if 'rapidcode' not in cfg:
        cfg['rapidcode'] = {}
    if 'axes' not in cfg['rapidcode']:
        cfg['rapidcode']['axes'] = {}

    for letter in ('A', 'B', 'C', 'D'):
        section = f'AXIS_{letter}'
        if section not in ini:
            continue
        s = ini[section]
        cfg['rapidcode']['axes'][letter] = {
            'name':              s.get('description', f'Axis {letter}'),
            'scaling':           float(s.get('scaling', 1.0)),
            'gearbox':           float(s.get('gearbox', 1.0)),
            'min_pos':           float(s.get('min', 0.0)),
            'max_pos':           float(s.get('max', 360.0)),
            'software_limit_deg': float(s.get('max', 360.0)),
            'speed_min_dps':     float(s.get('speed_min', 0.0)),
            'speed_max_dps':     float(s.get('speed_max', 360.0)),
            'accel_min_dps2':    float(s.get('accel_min', 0.0)),
            'accel_max_dps2':    float(s.get('accel_max', 720.0)),
            'jerk':              float(s.get('jerk', 30.0)),
        }

    # Sync Axis E into clearcore.axis_e
    # Note: Axis E is controlled directly by the GUI via UDP, not through TIM service.
    # The yaml sync keeps the service config consistent for reference/future routing.
    if 'AXIS_E' in ini:
        s = ini['AXIS_E']
        if 'clearcore' not in cfg:
            cfg['clearcore'] = {}
        cfg['clearcore']['axis_e'] = {
            'name':              s.get('description', 'Address Angle'),
            'min_pos':           float(s.get('min', 0.0)),
            'max_pos':           float(s.get('max', 45.0)),
            'software_limit_deg': float(s.get('max', 45.0)),
            'speed_min_dps':     float(s.get('speed_min', 0.0)),
            'speed_max_dps':     float(s.get('speed_max', 100.0)),
            'accel_min_dps2':    float(s.get('accel_min', 0.0)),
            'accel_max_dps2':    float(s.get('accel_max', 100.0)),
        }

    try:
        with open(yaml_path, 'w') as f:
            yaml.dump(cfg, f, default_flow_style=False, sort_keys=False)
        print(f'[INFO] Axis config synced from INI to {yaml_path}')
    except Exception as e:
        print(f'[WARNING] Could not write tim_config.yaml at {yaml_path}: {e}')


sync_ini_to_yaml(INI_PATH)
for axis in 'ABCDEFGH':
    section = f'AXIS_{axis}'
    if section in axis_ini:
        AXIS_UNITS[axis] = {
            'min': float(axis_ini[section]['min']),
            'max': float(axis_ini[section]['max']),
            'pulses': float(axis_ini[section]['pulses']),
            'degrees': float(axis_ini[section]['degrees']),
            'scaling': float(axis_ini[section]['scaling']),
            'gearbox': float(axis_ini[section]['gearbox']),
            'reverse': axis_ini[section].get('reverse', 'false').strip().lower() == 'true',
            'speed_min': float(axis_ini[section].get('speed_min', '0')),
            'speed_max': float(axis_ini[section].get('speed_max', '360')),
            'accel_min': float(axis_ini[section].get('accel_min', '0')),
            'accel_max': float(axis_ini[section].get('accel_max', '180')),
            'decel_min': float(axis_ini[section].get('decel_min', '0')),
            'decel_max': float(axis_ini[section].get('decel_max', '180')),
            'description': axis_ini[section].get('description', DEFAULT_SERVO_DESCRIPTIONS.get(ord(axis)-64, '')),
        }
# ============================================================================


# ============================================================================
#                         PLATFORM DETECTION & CONFIGURATION
# ============================================================================

# Cross-platform compatibility flags
IS_WINDOWS = platform.system() == "Windows"      # Windows development environment
IS_RASPBERRY_PI = platform.system() == "Linux" and platform.machine().startswith('arm')  # Pi deployment

GLOBAL_FONT = ('Courier New', 10)
POSITION_LABEL_FONT = ('Courier New', 10, 'bold')
CLEAR_BUTTON_FONT = ('Courier New', 9)
LOG_POSITION_POLLS = False  # Toggle to suppress noisy MG _RP logs in the GUI
DEFAULT_INPUT_BG = '#FFFFFF'
HIGHLIGHT_INPUT_BG = '#CCFFCC'
PENDING_INPUT_BG = '#FFFACD'


# ============================================================================
# Utility Functions
# ============================================================================
def format_display_value(val):
    """Display ints without decimals; otherwise one decimal place."""
    try:
        fval = round(float(val), 1)
        if fval.is_integer():
            return str(int(fval))
        return f"{fval:.1f}"
    except Exception:
        return str(val)


def pulses_to_degrees(raw_val, axis_letter):
    """Convert controller pulses to degrees using scaling and gearbox."""
    try:
        axis_units = AXIS_UNITS.get(axis_letter, {})
        scaling = axis_units.get('scaling', 1) or 1
        gearbox = axis_units.get('gearbox', 1) or 1
        return float(raw_val) / (scaling * gearbox)
    except Exception:
        return None


def bind_jog_press_release(window):
    """Bind mouse press/release for jog CW/CCW buttons to synthesize events."""
    # [CHANGE 2026-03-24 13:36:00 -04:00] Safety: disable press/release jog bindings to avoid hold-runaway behavior.
    return


def set_pending_highlight(window, servo_num, field):
    """Set field background to yellow if current value differs from last confirmed setpoint."""
    try:
        key = f'S{servo_num}_{field}'
        current = window[key].get()
        if not current or current in ('-', '.'):  # nothing meaningful
            window[key].update(background_color=DEFAULT_INPUT_BG)
            return False
        if not hasattr(window, '_last_setpoints') or not window._last_setpoints:
            window[key].update(background_color=PENDING_INPUT_BG)
            return True
        last_val = window._last_setpoints[servo_num - 1].get(field)
        try:
            current_val = float(current)
            pending = (last_val is None) or abs(current_val - float(last_val)) > 1e-6
        except Exception:
            pending = True
        window[key].update(background_color=PENDING_INPUT_BG if pending else DEFAULT_INPUT_BG)
        return pending
    except Exception:
        return False


def _axis_letter_for_index(idx: int) -> str:
    """Map 1-based axis index to Galil axis letter (A-H)."""
    letters = AXIS_LETTERS
    if 1 <= idx <= len(letters):
        return letters[idx - 1]
    return ''


def _clamp_and_convert_deg_to_pulses(axis_letter: str, deg_val: float) -> Dict[str, Any]:
    """Clamp degrees to axis limits, convert to pulses using scaling and gearbox."""
    axis_units = AXIS_UNITS.get(axis_letter, {})
    min_val = axis_units.get('min', NUMERIC_LIMITS['abs_pos'][0])
    max_val = axis_units.get('max', NUMERIC_LIMITS['abs_pos'][1])
    clamped_deg = max(min_val, min(max_val, deg_val))
    scaling = axis_units.get('scaling', 1) or 1
    gearbox = axis_units.get('gearbox', 1) or 1
    pulses = int(round(clamped_deg * scaling * gearbox))
    return {'deg': clamped_deg, 'pulses': pulses}


def save_axis_description(axis_letter: str, description: str):
    """Persist axis description to controller_config.ini for the given axis."""
    try:
        config = configparser.ConfigParser()
        config.read(INI_PATH)
        section = f'AXIS_{axis_letter}'
        if section not in config:
            return
        config[section]['description'] = description
        with open(INI_PATH, 'w') as fh:
            config.write(fh)
    except Exception:
        pass


def _refresh_description_colors(window):
    """Force description inputs to black text on white background (theme-safe)."""
    for i in range(1, 9):
        key = f'S{i}_desc'
        if key in window.AllKeysDict:
            try:
                window[key].update(text_color='black', background_color=DEFAULT_INPUT_BG)
                # Force underlying Tk widget colors to override dark themes
                widget = getattr(window[key], 'Widget', None)
                if widget is not None:
                    try:
                        widget.configure(fg='black', bg=DEFAULT_INPUT_BG, insertbackground='black')
                        # Some themes use readonlybackground; set defensively
                        widget.configure(readonlybackground=DEFAULT_INPUT_BG)
                    except Exception:
                        pass
            except Exception:
                pass


def load_datapipe_segments(file_path: str, sheet_name: str = None, row_start: int = 2, row_end: int = 61):
    """Load up to 60 segments from Excel with columns Time, Axis 1-5 (degrees).

    Header detection scans the first few rows (up to 10) to tolerate description rows above the real headers.
    Time is auto-interpreted: if values look like seconds (e.g., 0.05), they are converted to milliseconds.
    Returns (segments, seconds_to_ms_applied, missing_axes).
    """
    if not HAS_OPENPYXL:
        raise RuntimeError('openpyxl is required to read Excel files. Please install it.')
    if not os.path.exists(file_path):
        raise FileNotFoundError(f'File not found: {file_path}')

    def _norm(val):
        if val is None:
            return ''
        return ''.join(str(val).strip().lower().replace('_', ' ').split())

    required = ['time', 'axis1', 'axis2', 'axis3', 'axis4', 'axis5']
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    header_row = None
    headers = {}
    best_match = (-1, None, {})  # (score, row_idx, headers)
    # look for a header row in the first 10 rows to allow description rows above
    for row_idx in range(1, 11):
        row_headers = {}
        for cell in ws[row_idx]:
            norm = _norm(cell.value)
            if norm:
                row_headers[norm] = cell.column
        score = sum(1 for k in required if k in row_headers)
        if row_headers.get('time'):
            score += 0.5  # weight time a bit
        if score > best_match[0]:
            best_match = (score, row_idx, row_headers)
        if all(key in row_headers for key in required):
            header_row = row_idx
            headers = row_headers
            break
    if header_row is None:
        score, row_idx, row_headers = best_match
        if row_headers.get('time') and score >= 3.5:  # time + at least 3 axes
            header_row = row_idx
            headers = row_headers
        else:
            raise ValueError('Expected header row with: Time, Axis 1, Axis 2, Axis 3, Axis 4, Axis 5 (within first 10 rows).')

    time_col = headers.get('time')
    if time_col is None:
        raise ValueError('Expected column: Time')

    axis_cols = []
    missing_axes = []
    for i in range(1, 6):
        col = headers.get(f'axis{i}')
        axis_cols.append(col)
        if col is None:
            missing_axes.append(f'Axis {i}')

    data_start = max(row_start, header_row + 1)
    segments = []
    raw_times = []
    for row_idx in range(data_start, row_end + 1):
        time_cell = ws.cell(row=row_idx, column=time_col)
        time_val = time_cell.value
        if time_val in (None, ''):
            continue
        try:
            time_raw = float(time_val)
        except Exception:
            continue
        axis_vals = []
        for col in axis_cols:
            if col is None:
                axis_vals.append(0.0)
                continue
            val = ws.cell(row=row_idx, column=col).value
            axis_vals.append(float(val) if val not in (None, '') else 0.0)
        raw_times.append(time_raw)
        segments.append({'time_ms': time_raw, 'axis_deg': axis_vals})
        if len(segments) >= 60:
            break
    if not segments:
        return segments, False, missing_axes
    max_time = max(raw_times) if raw_times else 0
    min_time = min(raw_times) if raw_times else 0
    seconds_guess = (max_time <= 100) and (min_time < 10)
    if seconds_guess:
        for seg in segments:
            seg['time_ms'] = seg['time_ms'] * 1000.0
    return segments, seconds_guess, missing_axes


def build_datapipe_tab(default_path: str):
    """Build the DataPipe tab for loading Excel and sending contour data."""
    return [
        [sg.Text('Excel File'), sg.Input(default_path, key='DP_FILE', size=(60,1)), sg.FileBrowse('Browse', key='DP_BROWSE', target='DP_FILE')],
        [sg.Text('Sheet Name'), sg.Input('Galil', key='DP_SHEET', size=(12,1)),
         sg.Text('Rows'), sg.Input('2', key='DP_ROW_START', size=(4,1)), sg.Text('to'), sg.Input('61', key='DP_ROW_END', size=(4,1)),
         sg.Text('Time column assumed milliseconds')],
        [sg.Button('Load Preview', key='DP_LOAD', button_color=('white', 'green')),
         sg.Button('Send to Controller', key='DP_SEND', button_color=('white', '#007ACC'), disabled=True),
         sg.Button('Send via PR', key='DP_SEND_PR', button_color=('white', '#444444'), disabled=True),
         sg.Button('Send Batch PR', key='DP_SEND_BATCH_PR', button_color=('white', '#8800CC'), disabled=True),
         sg.Text('Run rows', size=(8,1)), sg.Input('5', key='DP_RUN_ROWS', size=(6,1), justification='center'),
         sg.Text('', key='DP_STATUS', size=(50,1))],
        [sg.Multiline('', key='DP_PREVIEW', size=(95,12), font=('Courier New', 9), autoscroll=True, write_only=True, disabled=False, border_width=1)]
    ]


def build_pvt_tab(default_path: str, default_sample_ms: int = 50):
    """Build the PVT tab for axes A-D with fixed sample time (default 50 ms)."""
    return [
        [sg.Text('File'), sg.Input(default_path, key='PVT_FILE', size=(60,1)), sg.FileBrowse('Browse', key='PVT_BROWSE', target='PVT_FILE')],
        [sg.Text('Sample (ms)', size=(12,1), font=GLOBAL_FONT),
         sg.Input(str(default_sample_ms), key='PVT_SAMPLE_MS', size=(6,1), justification='center'),
         sg.Text('Axes A-D · derived velocities', font=GLOBAL_FONT)],
        [sg.Button('Load Preview', key='PVT_LOAD', button_color=('white', 'green')),
         sg.Button('Send PVT', key='PVT_SEND', button_color=('white', '#007ACC'), disabled=True),
         sg.Text('', key='PVT_STATUS', size=(50,1))],
        [sg.Multiline('', key='PVT_PREVIEW', size=(95,12), font=('Courier New', 9), autoscroll=True, write_only=True, disabled=False, border_width=1)]
    ]


def _derive_velocities_deg(series_deg: List[float], sample_ms: float) -> List[float]:
    """Central/forward/backward difference velocities in deg/s."""
    dt = max(1e-6, sample_ms / 1000.0)
    n = len(series_deg)
    if n == 0:
        return []
    if n == 1:
        return [0.0]
    v = [0.0] * n
    v[0] = (series_deg[1] - series_deg[0]) / dt
    for i in range(1, n - 1):
        v[i] = (series_deg[i + 1] - series_deg[i - 1]) / (2 * dt)
    v[-1] = (series_deg[-1] - series_deg[-2]) / dt
    return v


def load_pvt_points(file_path: str, max_points: int = 200) -> List[List[float]]:
    """Load PVT position rows (deg) for axes A-D from CSV/XLSX. Missing cells default to 0."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f'File not found: {file_path}')
    ext = os.path.splitext(file_path)[1].lower()
    rows: List[List[float]] = []
    def _coerce(val):
        try:
            return float(val)
        except Exception:
            return None
    if ext in ('.xlsx', '.xlsm', '.xls'):
        if not HAS_OPENPYXL:
            raise RuntimeError('openpyxl is required to read Excel PVT files. Please install it.')
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            vals = [cell.value for cell in row[:4]]
            if all(v in (None, '') for v in vals):
                continue
            parsed = []
            for v in vals:
                c = _coerce(v)
                parsed.append(c if c is not None else 0.0)
            while len(parsed) < 4:
                parsed.append(0.0)
            if all(p == 0.0 for p in parsed):
                continue
            rows.append(parsed[:4])
            if len(rows) >= max_points:
                break
    else:
        with open(file_path, newline='') as fh:
            reader = csv.reader(fh)
            for line_idx, row in enumerate(reader):
                if not row:
                    continue
                # Skip header if first row is non-numeric
                if line_idx == 0:
                    numeric_present = any(_coerce(row[i]) is not None for i in range(min(4, len(row))))
                    if not numeric_present:
                        continue
                vals = row[:4]
                parsed = []
                for v in vals:
                    c = _coerce(v)
                    parsed.append(c if c is not None else 0.0)
                while len(parsed) < 4:
                    parsed.append(0.0)
                if all(p == 0.0 for p in parsed):
                    continue
                rows.append(parsed[:4])
                if len(rows) >= max_points:
                    break
    if not rows:
        raise ValueError('No PVT data rows found (expect positions for A-D).')
    return rows


def prepare_pvt_payload(rows_deg: List[List[float]], sample_ms: float):
    """Clamp to limits, convert to pulses, and derive velocities for axes A-D."""
    if sample_ms <= 0:
        raise ValueError('Sample time must be positive.')
    axes = AXIS_LETTERS[:4]
    count = len(rows_deg)
    axis_payload = {}
    for axis_idx, axis_letter in enumerate(axes):
        series_deg = []
        series_pulses = []
        for row in rows_deg:
            val_deg = float(row[axis_idx]) if axis_idx < len(row) else 0.0
            conv = _clamp_and_convert_deg_to_pulses(axis_letter, val_deg)
            series_deg.append(conv['deg'])
            series_pulses.append(conv['pulses'])
        vel_deg = _derive_velocities_deg(series_deg, sample_ms)
        axis_units = AXIS_UNITS.get(axis_letter, {})
        scale = axis_units.get('scaling', 1) or 1
        gearbox = axis_units.get('gearbox', 1) or 1
        vel_pulses = [int(round(v * scale * gearbox)) for v in vel_deg]
        axis_payload[axis_letter] = {
            'deg': series_deg,
            'pulses': series_pulses,
            'vel_deg': vel_deg,
            'vel_pulses': vel_pulses,
        }
    return {'sample_ms': sample_ms, 'count': count, 'axis': axis_payload}


def render_pvt_preview(window, payload):
    if not payload:
        window['PVT_PREVIEW'].update('')
        return
    sample_ms = payload['sample_ms']
    count = payload['count']
    lines = [f'Points: {count}   Sample: {sample_ms:.1f} ms   Axes: A-D']
    header = 'Idx  ' + '  '.join([f"{ax}:deg/pul  v(deg/s)" for ax in AXIS_LETTERS[:4]])
    lines.append(header)
    for idx in range(count):
        parts = [f"{idx+1:02d}"]
        for axis_letter in AXIS_LETTERS[:4]:
            axis_data = payload['axis'][axis_letter]
            deg_val = axis_data['deg'][idx]
            pulse_val = axis_data['pulses'][idx]
            vel_val = axis_data['vel_deg'][idx]
            parts.append(f"{axis_letter}:{deg_val:7.2f}/{pulse_val:7d} v:{vel_val:7.1f}")
        lines.append('  '.join(parts))
        if len(lines) >= 60:  # avoid overly tall preview
            lines.append('...')
            break
    window['PVT_PREVIEW'].update('\n'.join(lines))


def send_pvt_payload(comm, payload, window=None):
    """Send PVT arrays for axes A-D using PT/PV/PA with uniform sample time."""
    if comm is None:
        raise RuntimeError('Controller communications not initialized.')
    if getattr(comm, 'mode', None) != 'CommMode1':
        raise RuntimeError('PVT send only supported in CommMode1 (Galil Ethernet).')
    if not payload or not payload.get('axis'):
        raise ValueError('No PVT payload to send.')
    sample_ms = payload['sample_ms']
    count = payload['count']
    if count <= 0:
        raise ValueError('No PVT points to send.')
    cmds = ['ST']
    for axis_letter in AXIS_LETTERS[:4]:
        axis_data = payload['axis'].get(axis_letter)
        if not axis_data:
            continue
        time_csv = ','.join(str(int(round(sample_ms))) for _ in range(count))
        pos_csv = ','.join(str(int(p)) for p in axis_data['pulses'])
        vel_csv = ','.join(str(int(v)) for v in axis_data['vel_pulses'])
        cmds.append(f'PT{axis_letter}={time_csv}')
        cmds.append(f'PV{axis_letter}={vel_csv}')
        cmds.append(f'PA{axis_letter}={pos_csv}')
    cmds.append('BGS')
    for c in cmds:
        result = comm.send_command(c)
        if window and 'DEBUG_LOG' in window.AllKeysDict:
            window['DEBUG_LOG'].print(f"[PVT_CMD] {c} -> {result}")
        if result is False:
            raise RuntimeError(f'Controller rejected command: {c}')


def build_all_pvt_payload(values, window, sample_ms: float):
    """Build PVT rows from ALL tab setpoints (axes A-D) using current hold positions when blanks/disabled."""
    if sample_ms <= 0:
        raise ValueError('Sample time must be positive.')
    # Seed hold positions from last valid polled positions if available; fallback to 0
    hold_positions = {}
    for idx, axis_letter in enumerate(AXIS_LETTERS[:4], start=1):
        try:
            pos_str = window._last_valid_pos[idx - 1] if hasattr(window, '_last_valid_pos') else ''
            hold_positions[axis_letter] = float(pos_str) if pos_str not in (None, '', '-') else 0.0
        except Exception:
            hold_positions[axis_letter] = 0.0

    rows = []
    step_fields = ['pos1', 'pos2', 'pos3', 'pos4', 'pos5']
    for step_idx, pos_field in enumerate(step_fields, start=1):
        row = []
        any_axis = False
        for axis_idx, axis_letter in enumerate(AXIS_LETTERS[:4], start=1):
            enabled = bool(values.get(f'ALL_S{axis_idx}_enabled', False))
            key = f'ALL_S{axis_idx}_{pos_field}'
            raw = str(values.get(key, '')).strip()
            if enabled and raw not in ('', '-', '.', '-.'):
                try:
                    val = float(raw)
                    hold_positions[axis_letter] = val
                    any_axis = True
                except Exception:
                    raise ValueError(f'Invalid value for Servo {axis_idx} {pos_field}: {raw}')
            row.append(hold_positions[axis_letter])
        if any_axis:
            rows.append(row)

    if not rows:
        raise ValueError('No valid setpoints to send as PVT (provide pos1-pos4 for at least one of axes A-D).')

    return prepare_pvt_payload(rows, sample_ms)


# ---------------------------------------------------------------------------
# Sequence state persistence
# ---------------------------------------------------------------------------
def load_sequence_state():
    try:
        if os.path.exists(SEQUENCE_STATE_FILE):
            with open(SEQUENCE_STATE_FILE, 'r') as fh:
                return json.load(fh)
    except Exception:
        pass
    return {}


def save_sequence_state_from_values(values):
    state = {
        'repeat': bool(values.get('ALL_REPEAT', False)),
        'servos': {},
        'dp_run_rows': values.get('DP_RUN_ROWS', ''),
    }
    for i in range(1, 9):
        entry = {
            'enabled': bool(values.get(f'ALL_S{i}_enabled', False)),
            'pos1': values.get(f'ALL_S{i}_pos1', ''),
            'pos2': values.get(f'ALL_S{i}_pos2', ''),
            'pos3': values.get(f'ALL_S{i}_pos3', ''),
            'pos4': values.get(f'ALL_S{i}_pos4', ''),
            'pos5': values.get(f'ALL_S{i}_pos5', ''),
        }
        state['servos'][str(i)] = entry
    try:
        with open(SEQUENCE_STATE_FILE, 'w') as fh:
            json.dump(state, fh)
    except Exception:
        pass


def restore_sequence_state(window, state):
    try:
        if not state:
            return
        window['ALL_REPEAT'].update(value=bool(state.get('repeat', False)))
        # Restore DP run rows if present
        dp_rows = state.get('dp_run_rows', '')
        if dp_rows is not None and 'DP_RUN_ROWS' in window.AllKeysDict:
            window['DP_RUN_ROWS'].update(str(dp_rows))
        servos = state.get('servos', {})
        for i_str, entry in servos.items():
            try:
                i = int(i_str)
            except Exception:
                continue
            if f'ALL_S{i}_enabled' in window.AllKeysDict:
                window[f'ALL_S{i}_enabled'].update(value=bool(entry.get('enabled', False)))
            for fld in ['pos1', 'pos2', 'pos3', 'pos4', 'pos5']:
                key = f'ALL_S{i}_{fld}'
                if key in window.AllKeysDict:
                    val = entry.get(fld, '')
                    window[key].update(str(val))
    except Exception:
        pass


def load_motion_defaults():
    """Load persisted startup defaults for servo motion fields."""
    try:
        if os.path.exists(MOTION_DEFAULTS_FILE):
            with open(MOTION_DEFAULTS_FILE, 'r') as fh:
                data = json.load(fh)
                return data if isinstance(data, dict) else {}
    except Exception:
        pass
    return {}


def save_motion_defaults_from_values(values):
    """Persist speed/accel/decel values so they can be restored on next startup."""
    state = {'servos': {}}
    for i in range(1, 9):
        speed_raw = str(values.get(f'S{i}_speed', '')).strip()
        accel_raw = str(values.get(f'S{i}_accel', '')).strip()
        decel_raw = str(values.get(f'S{i}_decel', '')).strip()
        entry = {}
        try:
            if speed_raw not in ('', '-', '.', '-.'):
                entry['speed'] = float(speed_raw)
        except Exception:
            pass
        try:
            if accel_raw not in ('', '-', '.', '-.'):
                entry['accel'] = float(accel_raw)
        except Exception:
            pass
        try:
            if decel_raw not in ('', '-', '.', '-.'):
                entry['decel'] = float(decel_raw)
        except Exception:
            pass
        if entry:
            state['servos'][str(i)] = entry
    try:
        with open(MOTION_DEFAULTS_FILE, 'w') as fh:
            json.dump(state, fh)
    except Exception:
        pass


# [CHANGE 2026-03-24 11:29:00 -04:00] Rename startup defaults helper to reflect speed/accel/decel scope.
def apply_startup_motion_defaults(window):
    """Apply remembered speed/accel/decel or fallback defaults at program startup."""
    remembered = load_motion_defaults().get('servos', {})
    if not hasattr(window, '_last_setpoints') or not window._last_setpoints or len(window._last_setpoints) != 8:
        window._last_setpoints = [{f: None for f in ['speed', 'accel', 'decel', 'abs_pos', 'rel_pos', 'jog_amount']} for _ in range(8)]

    def _is_blank_or_zero(raw):
        s = str(raw).strip()
        if s in ('', '-', '.', '-.'):
            return True
        try:
            return abs(float(s)) < 1e-9
        except Exception:
            return False

    for i in range(1, 9):
        rem = remembered.get(str(i), {}) if isinstance(remembered, dict) else {}
        speed_val = rem.get('speed', STARTUP_SPEED_DEFAULT)
        accel_val = rem.get('accel', STARTUP_ACCEL_DECEL_DEFAULT)
        decel_val = rem.get('decel', STARTUP_ACCEL_DECEL_DEFAULT)

        try:
            speed_fmt = format_display_value(float(speed_val))
        except Exception:
            speed_fmt = format_display_value(STARTUP_SPEED_DEFAULT)
        try:
            accel_fmt = format_display_value(float(accel_val))
        except Exception:
            accel_fmt = format_display_value(STARTUP_ACCEL_DECEL_DEFAULT)
        try:
            decel_fmt = format_display_value(float(decel_val))
        except Exception:
            decel_fmt = format_display_value(STARTUP_ACCEL_DECEL_DEFAULT)

        speed_key = f'S{i}_speed'
        accel_key = f'S{i}_accel'
        decel_key = f'S{i}_decel'

        has_remembered_speed = isinstance(rem, dict) and ('speed' in rem)
        has_remembered_accel = isinstance(rem, dict) and ('accel' in rem)
        has_remembered_decel = isinstance(rem, dict) and ('decel' in rem)

        if speed_key in window.AllKeysDict:
            current_speed = str(window[speed_key].get()).strip()
            if has_remembered_speed or _is_blank_or_zero(current_speed):
                window[speed_key].update(speed_fmt)
        if accel_key in window.AllKeysDict:
            current_accel = str(window[accel_key].get()).strip()
            if has_remembered_accel or _is_blank_or_zero(current_accel):
                window[accel_key].update(accel_fmt)
        if decel_key in window.AllKeysDict:
            current_decel = str(window[decel_key].get()).strip()
            if has_remembered_decel or _is_blank_or_zero(current_decel):
                window[decel_key].update(decel_fmt)

        try:
            window._last_setpoints[i - 1]['speed'] = float(speed_fmt)
        except Exception:
            window._last_setpoints[i - 1]['speed'] = STARTUP_SPEED_DEFAULT
        try:
            window._last_setpoints[i - 1]['accel'] = float(accel_fmt)
        except Exception:
            window._last_setpoints[i - 1]['accel'] = STARTUP_ACCEL_DECEL_DEFAULT
        try:
            window._last_setpoints[i - 1]['decel'] = float(decel_fmt)
        except Exception:
            window._last_setpoints[i - 1]['decel'] = STARTUP_ACCEL_DECEL_DEFAULT

    # [CHANGE 2026-03-24 11:19:00 -04:00] Recompute midpoint speed labels after startup defaults are applied.
    for i in range(1, 9):
        update_mid_speed_display(window, i)
# -----------------------------
# Servo setup:Dictionary mapping each field to its min and max values (same for all servos)
# Automatically generate command mappings for all 8 servos
# The min an max fields for each servo is derived from AXIS_UNITS which are defined in the controller_config.ini file.
# -----------------------------
# Command Mapping Dictionaries
# -----------------------------
# GALIL_COMMAND_MAP: Maps GUI actions to Galil controller commands (A-H axes)
GALIL_COMMAND_MAP = {}
AXIS_LETTERS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
for i in range(1, 9):
    axis_letter = AXIS_LETTERS[i-1]
    GALIL_COMMAND_MAP[f'S{i}_enable'] = f'SH{axis_letter}'
    GALIL_COMMAND_MAP[f'S{i}_disable'] = f'MO{axis_letter}'
    GALIL_COMMAND_MAP[f'S{i}_clear_faults'] = f'CF{axis_letter}'
    GALIL_COMMAND_MAP[f'S{i}_start'] = f'BG{axis_letter}'
    GALIL_COMMAND_MAP[f'S{i}_stop'] = f'ST{axis_letter}'
    GALIL_COMMAND_MAP[f'S{i}_jog'] = (lambda speed, axis=axis_letter: f'JG{axis}={speed};BG{axis}')
    GALIL_COMMAND_MAP[f'S{i}_speed'] = (lambda value, axis=axis_letter: f'SP{axis}={value}')
    GALIL_COMMAND_MAP[f'S{i}_accel'] = (lambda value, axis=axis_letter: f'AC{axis}={value}')
    GALIL_COMMAND_MAP[f'S{i}_decel'] = (lambda value, axis=axis_letter: f'DC{axis}={value}')
    GALIL_COMMAND_MAP[f'S{i}_abs_pos'] = (lambda value, axis=axis_letter: f'PA{axis}={value}')
    GALIL_COMMAND_MAP[f'S{i}_rel_pos'] = (lambda value, axis=axis_letter: f'PR{axis}={value}')

# Galil only
COMMAND_MAP = GALIL_COMMAND_MAP

# Initialize ControllerComm with RSI and ClearCore from INI file
# Initialize ControllerComm with RSI and ClearCore from INI file
import os
comm = None       # RSI Software (axes A-D) - Servos 1-4
comm_e = None     # ClearCore (axis E) - Servo 5
comm_h = None     # MyActuator (axis H) - Servo 8

try:
    config = configparser.ConfigParser()
    ini_path = os.path.join(os.path.dirname(__file__), 'controller_config.ini')
    print(f'[DEBUG] Reading INI file from: {ini_path}')
    config.read(ini_path)
    print(f'[DEBUG] Sections found: {config.sections()}')
    
    # Initialize RSI for axes A-D (Servos 1-4) - OPTIONAL if not running yet
    if config.has_section('CommMode1'):
        rsi_config = dict(config.items('CommMode1'))
        print(f'[DEBUG] rsi_config: {rsi_config}')
        try:
            comm = ControllerComm(mode='CommMode1', rsi_config=rsi_config)
            print('[DEBUG] RSI Software initialized for Axes A-D (Servos 1-4)')
        except Exception as rsi_error:
            print(f'[WARNING] RSI not available: {rsi_error}')
            print('[INFO] Continuing without RSI - only Axis E (ClearCore) will be available')
            comm = None
    
    # Initialize ClearCore for axis E (Servo 5) - THIS WILL WORK INDEPENDENTLY
    if config.has_section('CommMode6'):
        clearcore_config = dict(config.items('CommMode6'))
        print(f'[DEBUG] clearcore_config: {clearcore_config}')
        comm_e = ControllerComm(mode='CommMode6', clearcore_config=clearcore_config)
        print('[DEBUG] ClearCore Board 1 initialized for Axis E (Servo 5)')
    
    # Initialize MyActuator for axis H (Servo 8) - OPTIONAL if not connected
    if config.has_section('CommMode5'):
        myactuator_config = dict(config.items('CommMode5'))
        print(f'[DEBUG] myactuator_config: {myactuator_config}')
        try:
            comm_h = ControllerComm(mode='CommMode5', myactuator_config=myactuator_config)
            print('[DEBUG] MyActuator initialized for Axis H (Servo 8)')
        except Exception as myact_error:
            print(f'[WARNING] MyActuator not available: {myact_error}')
            comm_h = None

except Exception as comm_error:
    comm = None
    comm_e = None
    comm_h = None
    import traceback
    error_details = f'{comm_error}\n' + traceback.format_exc()
    sg.popup_error(f'Failed to initialize controller communications:\n{comm_error}', keep_on_top=True)
    print(f'[ERROR] Failed to initialize controller communications: {error_details}')


def build_servo_tab(servo_num):
###############################################################################
    """
    Build the tab layout for a single servo.
    Includes status indicator, value fields, and control buttons.
    Args:
        servo_num (int): Servo number (1-8)
    Returns:
        List: PySimpleGUI layout for the tab
    """
    axis_letter = AXIS_LETTERS[servo_num - 1]
    desc_default = AXIS_UNITS.get(axis_letter, {}).get('description', DEFAULT_SERVO_DESCRIPTIONS.get(servo_num, ''))
    layout = [
        [sg.Text(f'Servo {servo_num}', font=POSITION_LABEL_FONT),
         sg.Text('●', key=f'S{servo_num}_status_light', font=('Courier New', 16), text_color='gray'),
         sg.Text('Disabled', key=f'S{servo_num}_status_text', font=GLOBAL_FONT, text_color='gray')],
        # [CHANGE 2026-04-17 00:00:00 -04:00] Comm health indicator: shows link status for this axis's controller.
        [sg.Text('Link:', font=GLOBAL_FONT),
         sg.Text('●', key=f'S{servo_num}_comm_indicator', font=('Courier New', 14), text_color='gray',
             tooltip='Controller communications health (green=OK, red=failed, gray=not configured)'),
         sg.Text('—', key=f'S{servo_num}_comm_label', font=GLOBAL_FONT, text_color='gray', size=(12,1)),
         sg.Button('Reconnect', key=f'S{servo_num}_reconnect', font=GLOBAL_FONT, size=(10,1),
             button_color=('white', '#444444'), tooltip='Re-establish controller communications link')],
        [sg.Text('Description:', size=(18,1), font=GLOBAL_FONT),
         sg.Input(
             default_text=desc_default,
             key=f'S{servo_num}_desc',
             size=(24,1),
             font=GLOBAL_FONT,
             enable_events=True,
             text_color='black',
             background_color=DEFAULT_INPUT_BG,
         )],
        [
            sg.Button('Clear Faults', key=f'S{servo_num}_clear_faults', size=(14,2), font=GLOBAL_FONT),
            sg.Checkbox(
                'Confirm setpoints',
                key=f'S{servo_num}_confirm_ok',
                default=True,
                font=GLOBAL_FONT,
                tooltip='When checked, show OK popup before sending setpoints.',
            ),
        ],
          [sg.Text('Speed:', size=(18,1), font=GLOBAL_FONT),
            sg.Input('10', key=f'S{servo_num}_speed', size=(10,1), font=GLOBAL_FONT, enable_events=True),
                sg.Text('DPS ', font=GLOBAL_FONT),
            sg.Button('⌨', key=f'S{servo_num}_speed_keypad', size=(2,1), font=GLOBAL_FONT),
            sg.Button('OK', key=f'S{servo_num}_speed_ok', size=(4,1), font=GLOBAL_FONT, button_color=('white', 'green')),
            sg.Text(' @mid:', font=GLOBAL_FONT, pad=((10,2),(0,0))),
            sg.Text('—', key=f'S{servo_num}_mid_speed', size=(10,1), font=GLOBAL_FONT, text_color='blue'),
            sg.Text('DPS', font=GLOBAL_FONT)],
        [sg.Text('Acceleration:', size=(18,1), font=GLOBAL_FONT),
         sg.Input(key=f'S{servo_num}_accel', size=(10,1), font=GLOBAL_FONT, enable_events=True),
            sg.Text('DPS²', font=GLOBAL_FONT),
         sg.Button('⌨', key=f'S{servo_num}_accel_keypad', size=(2,1), font=GLOBAL_FONT),
         sg.Button('OK', key=f'S{servo_num}_accel_ok', size=(4,1), font=GLOBAL_FONT, button_color=('white', 'green'))],
        [sg.Text('Deceleration:', size=(18,1), font=GLOBAL_FONT),
         sg.Input(key=f'S{servo_num}_decel', size=(10,1), font=GLOBAL_FONT, enable_events=True),
            sg.Text('DPS²', font=GLOBAL_FONT),
         sg.Button('⌨', key=f'S{servo_num}_decel_keypad', size=(2,1), font=GLOBAL_FONT),
         sg.Button('OK', key=f'S{servo_num}_decel_ok', size=(4,1), font=GLOBAL_FONT, button_color=('white', 'green'))],
        [sg.Text('Absolute Position (DEG):', size=(18,1), font=GLOBAL_FONT),
            sg.Input(key=f'S{servo_num}_abs_pos', size=(10,1), font=GLOBAL_FONT, enable_events=True),
                sg.Text('DEG ', font=GLOBAL_FONT),
            sg.Button('⌨', key=f'S{servo_num}_abs_pos_keypad', size=(2,1), font=GLOBAL_FONT),
            sg.Button('OK', key=f'S{servo_num}_abs_pos_ok', size=(4,1), font=GLOBAL_FONT, button_color=('white', 'green'))],
        [sg.Text('Relative Position (DEG):', size=(18,1), font=GLOBAL_FONT),
            sg.Input(key=f'S{servo_num}_rel_pos', size=(10,1), font=GLOBAL_FONT, enable_events=True),
                sg.Text('DEG ', font=GLOBAL_FONT),
            sg.Button('⌨', key=f'S{servo_num}_rel_pos_keypad', size=(2,1), font=GLOBAL_FONT),
            sg.Button('OK', key=f'S{servo_num}_rel_pos_ok', size=(4,1), font=GLOBAL_FONT, button_color=('white', 'green'))],
        [sg.Text('Actual Position:', size=(18,1), font=GLOBAL_FONT), sg.Text('0', key=f'S{servo_num}_actual_pos', size=(10,1), font=GLOBAL_FONT), sg.Text('DEG', font=GLOBAL_FONT)],
        [sg.Text('Actual Position:', size=(18,1), font=GLOBAL_FONT), sg.Text('0', key=f'S{servo_num}_actual_pos_pulses', size=(10,1), font=GLOBAL_FONT), sg.Text('PUL', font=GLOBAL_FONT)],
        [
            sg.Button('Servo Enable', key=f'S{servo_num}_enable', size=(12,2), font=GLOBAL_FONT),
            sg.Button('Servo Disable', key=f'S{servo_num}_disable', size=(12,2), font=GLOBAL_FONT),
            sg.Button('Zero Position', key=f'S{servo_num}_zero_pos', size=(12,2), font=GLOBAL_FONT)  # Default color (same as Jog)
        ],
        [
            sg.Button('Start Motion', key=f'S{servo_num}_start', size=(12,2), font=GLOBAL_FONT),
            sg.Button('Stop Motion', key=f'S{servo_num}_stop', size=(12,2), font=GLOBAL_FONT),
            sg.Button('Jog CW', key=f'S{servo_num}_jog_cw', size=(8,2), font=GLOBAL_FONT),
            sg.Button('Jog CCW', key=f'S{servo_num}_jog_ccw', size=(8,2), font=GLOBAL_FONT),
            sg.Text('Jog Amt (deg):', font=GLOBAL_FONT),
            sg.Input(
                default_text=format_display_value(JOG_STEP_DEFAULT_DEG),
                key=f'S{servo_num}_jog_amount',
                size=(5,1),
                font=GLOBAL_FONT,
                enable_events=True,
                justification='center'
            ),
            sg.Button('⌨', key=f'S{servo_num}_jog_amount_keypad', size=(2,1), font=GLOBAL_FONT)
        ]
    ]
    return layout


def build_all_tab():
    """Build the ALL tab with up to four absolute setpoints per servo and enable checkbox."""
    rows = [
        [
            sg.Text('Line Speed (0-2):', size=(18, 1), font=GLOBAL_FONT),
            sg.Input('1.0', key='ALL_LINE_SPEED', size=(6, 1), font=GLOBAL_FONT, justification='center', enable_events=True),
        ],
        [
            sg.Text('PVT Sample (ms):', size=(18, 1), font=GLOBAL_FONT),
            sg.Input('50', key='ALL_PVT_SAMPLE_MS', size=(6, 1), font=GLOBAL_FONT, justification='center'),
        ],
        [
            sg.Text('Enable', size=(6, 1), font=GLOBAL_FONT, justification='center'),
            sg.Text('Description', size=(21, 1), font=GLOBAL_FONT, justification='center'),
            sg.Text('SPT 1(DEG)', size=(10, 1), font=GLOBAL_FONT, justification='center'),
            sg.Text('SPT 2(DEG)', size=(10, 1), font=GLOBAL_FONT, justification='center'),
            sg.Text('SPT 3(DEG)', size=(10, 1), font=GLOBAL_FONT, justification='center'),
            sg.Text('SPT 4(DEG)', size=(10, 1), font=GLOBAL_FONT, justification='center'),
            sg.Text('SPT 5(DEG)', size=(10, 1), font=GLOBAL_FONT, justification='center')
        ]
    ]
    rows.append([
        sg.Text('', size=(6, 1), font=GLOBAL_FONT),
        sg.Text('Cycle Time (s)', size=(21, 1), font=GLOBAL_FONT, justification='center'),
        sg.Text('—', key='ALL_STEP1_TIME', size=(10, 1), font=GLOBAL_FONT, justification='center'),
        sg.Text('—', key='ALL_STEP2_TIME', size=(10, 1), font=GLOBAL_FONT, justification='center'),
        sg.Text('—', key='ALL_STEP3_TIME', size=(10, 1), font=GLOBAL_FONT, justification='center'),
        sg.Text('—', key='ALL_STEP4_TIME', size=(10, 1), font=GLOBAL_FONT, justification='center'),
        sg.Text('—', key='ALL_STEP5_TIME', size=(10, 1), font=GLOBAL_FONT, justification='center'),
    ])
    for i in range(1, 9):
        axis_letter = AXIS_LETTERS[i - 1]
        desc_default = AXIS_UNITS.get(axis_letter, {}).get('description', DEFAULT_SERVO_DESCRIPTIONS.get(i, ''))
        rows.append([
            sg.Checkbox(f'Servo {i}', key=f'ALL_S{i}_enabled', default=True, font=GLOBAL_FONT, pad=((0,4),0)),
            sg.Text(desc_default, key=f'ALL_S{i}_desc', size=(18, 1), font=GLOBAL_FONT, pad=((0,4),0)),
            sg.Input(key=f'ALL_S{i}_pos1', size=(10, 1), font=GLOBAL_FONT, enable_events=True, justification='center'),
            sg.Input(key=f'ALL_S{i}_pos2', size=(10, 1), font=GLOBAL_FONT, enable_events=True, justification='center'),
            sg.Input(key=f'ALL_S{i}_pos3', size=(10, 1), font=GLOBAL_FONT, enable_events=True, justification='center'),
            sg.Input(key=f'ALL_S{i}_pos4', size=(10, 1), font=GLOBAL_FONT, enable_events=True, justification='center'),
            sg.Input(key=f'ALL_S{i}_pos5', size=(10, 1), font=GLOBAL_FONT, enable_events=True, justification='center'),
        ])
    rows.append([
        sg.Button('Run Sequence', key='ALL_RUN_SEQUENCE', size=(14, 2), font=GLOBAL_FONT, button_color=('white', 'green')),
        sg.Button('Stop Sequence', key='ALL_STOP_SEQUENCE', size=(14, 2), font=GLOBAL_FONT, button_color=('white', 'red'), disabled=True),
        sg.Checkbox('Repeat', key='ALL_REPEAT', default=False, font=GLOBAL_FONT)
    ])
    rows.append([
        sg.Button('Send as PVT', key='ALL_PVT_SEND', size=(14, 2), font=GLOBAL_FONT, button_color=('white', '#007ACC')),
        sg.Text('', key='ALL_PVT_STATUS', size=(40, 1), font=GLOBAL_FONT),
    ])
    return rows



# Create a tab for each servo (1-8)
# -----------------------------
# Build tab group for all servos
# -----------------------------
TAB_PAD = ((6, 6), (0, 0))  # consistent tab header padding
servo_tabs = [sg.Tab(f'Servo {i}', build_servo_tab(i), key=f'TAB{i}', pad=TAB_PAD) for i in range(1, 9)]
servo_tabs.append(sg.Tab(' ALL ', build_all_tab(), key='TABALL', pad=TAB_PAD))
servo_tabs.append(sg.Tab(' DataPipe ', build_datapipe_tab(os.path.join(os.path.dirname(__file__), 'MatlabData5.xlsx')), key='TABDP', pad=TAB_PAD))
servo_tabs.append(sg.Tab(' PVT ', build_pvt_tab(os.path.join(os.path.dirname(__file__), 'pvt_points.csv')), key='TABPVT', pad=TAB_PAD))



# List of all numeric input keys and keypad button keys for all servos
# -----------------------------
# Build lists of input keys and keypad button keys for all servos
# -----------------------------

# Default numeric limits for fields (used if not found in AXIS_UNITS)
NUMERIC_LIMITS = {
    # Allow 0–360 deg/sec for speed; direction is handled by command sign
    'speed': (0, 360),
    'accel': (0, 180),
    'decel': (0, 180),
    'abs_pos': (0, 180),
    'rel_pos': (-90, 90),
    'jog_amount': (JOG_STEP_MIN_DEG, JOG_STEP_MAX_DEG),
}


def get_limits(axis_letter: str, field: str):
    """Return (min, max) for a field, preferring axis-specific overrides."""
    axis_cfg = AXIS_UNITS.get(axis_letter, {})
    default_min, default_max = NUMERIC_LIMITS.get(field, (None, None))
    if field == 'rel_pos':
        custom_abs_min = axis_cfg.get('min')
        custom_abs_max = axis_cfg.get('max')
        if custom_abs_min is not None and custom_abs_max is not None:
            rel_span = abs(float(custom_abs_max) - float(custom_abs_min))
            if rel_span > 0:
                return -rel_span, rel_span
        return default_min, default_max
    if field == 'abs_pos':
        custom_min = axis_cfg.get('min')
        custom_max = axis_cfg.get('max')
    else:
        custom_min = axis_cfg.get(f'{field}_min')
        custom_max = axis_cfg.get(f'{field}_max')
    min_val = custom_min if custom_min is not None else default_min
    max_val = custom_max if custom_max is not None else default_max
    return min_val, max_val


def compute_midpoint_speed(speed: float, accel: float, decel: float, mid_distance: float) -> float | None:
    """Estimate speed achievable at the midpoint of a move.

    Uses a trapezoidal/triangular heuristic: the achievable speed at midpoint is limited by
    commanded speed and by accel/decel over the available distance to halfway.
    """
    try:
        if any(v is None for v in [speed, accel, decel]):
            return None
        speed = max(0.0, float(speed))
        accel = max(0.0, float(accel))
        decel = max(0.0, float(decel))
        mid_distance = max(0.0, float(mid_distance))
        if mid_distance == 0 or accel == 0 or decel == 0:
            return 0.0
        limit_accel = (2 * accel * mid_distance) ** 0.5
        limit_decel = (2 * decel * mid_distance) ** 0.5
        return min(speed, limit_accel, limit_decel)
    except Exception:
        return None


def update_mid_speed_display(window, servo_num: int):
    """Update the displayed midpoint speed estimate for a servo."""
    try:
        axis_letter = AXIS_LETTERS[servo_num - 1]
        axis_cfg = AXIS_UNITS.get(axis_letter, {})
        axis_range = (axis_cfg.get('max', 180.0) - axis_cfg.get('min', 0.0))
        # Use half of axis range but cap at 90 deg as requested
        mid_distance = max(0.0, min(90.0, axis_range / 2.0))
        def _parse(key):
            val = window[key].get() if key in window.AllKeysDict else None
            try:
                return float(val)
            except Exception:
                return None
        speed = _parse(f'S{servo_num}_speed')
        accel = _parse(f'S{servo_num}_accel')
        decel = _parse(f'S{servo_num}_decel')
        mid_speed = compute_midpoint_speed(speed, accel, decel, mid_distance)
        display_key = f'S{servo_num}_mid_speed'
        if display_key in window.AllKeysDict:
            if mid_speed is None:
                window[display_key].update('—')
            else:
                window[display_key].update(f"{format_display_value(mid_speed)} DPS")
    except Exception:
        try:
            window[f'S{servo_num}_mid_speed'].update('—')
        except Exception:
            pass

# -----------------------------
# Main window layout
# -----------------------------
NUMERIC_INPUT_KEYS = []
NUMERIC_KEYPAD_BUTTONS = []
for i in range(1, 9):
    for field in ['speed', 'accel', 'decel', 'abs_pos', 'rel_pos', 'jog_amount']:
        NUMERIC_INPUT_KEYS.append(f'S{i}_{field}')
        NUMERIC_KEYPAD_BUTTONS.append(f'S{i}_{field}_keypad')
    for pos_field in ['pos1', 'pos2', 'pos3', 'pos4', 'pos5']:
        NUMERIC_INPUT_KEYS.append(f'ALL_S{i}_{pos_field}')

# DataPipe helpers tracked on window
DP_SEGMENTS_KEY = '_dp_segments'
DP_TIME_KEY = '_dp_time_ms'

###############################################################################
# GUI Layout
###############################################################################
# Make DEBUG_LOG Multiline write-only so updates render but user can't edit
layout = [
    [
        sg.TabGroup([servo_tabs], key='TABGROUP', enable_events=True, expand_x=True, expand_y=True)
    ],
    [
        sg.Column(
            [[sg.Multiline('', key='DEBUG_LOG', size=(55, 8), font=('Courier New', 9), autoscroll=True, disabled=False, write_only=True, text_color='black', border_width=0)]],
            expand_x=True,
            expand_y=True,
        ),
        sg.Column(
            [
                [
                    sg.Checkbox('Show Poll Logs', key='SHOW_POLL_LOGS', enable_events=True, default=False, font=GLOBAL_FONT),
                    sg.Button('Shutdown', key='SHUTDOWN', size=(10,2), button_color=('white', 'red'), font=GLOBAL_FONT),
                    sg.Button('E-STOP', key='ESTOP', size=(10,2), button_color=('white', '#C00000'), font=('Courier New', 10, 'bold'))
                ]
            ],
            element_justification='right',
            pad=((10,0), (0,0)),
            vertical_alignment='top'
        )
    ]
]

window = sg.Window("Controller GUI", layout, size=(800, 540), font=GLOBAL_FONT, finalize=True, return_keyboard_events=True, resizable=True)

# Route WARNING+ log records to the GUI debug log so comm errors are visible
# without a terminal session open.
class _GUILogHandler(logging.Handler):
    def __init__(self, win):
        super().__init__(level=logging.WARNING)
        self._win = win
    def emit(self, record):
        try:
            self._win.write_event_value('GUI_LOG', self.format(record))
        except Exception:
            pass

_gui_log_handler = _GUILogHandler(window)
_gui_log_handler.setFormatter(logging.Formatter('%(levelname)s %(name)s: %(message)s'))
logging.getLogger().addHandler(_gui_log_handler)

# Enforce description field colors after window creation
_refresh_description_colors(window)

# Restore saved ALL tab sequence state (repeat flag, enable flags, setpoints)
restore_sequence_state(window, load_sequence_state())

# [CHANGE 2026-03-24 11:19:00 -04:00] Apply remembered speed/accel/decel startup defaults (fallback=10).
apply_startup_motion_defaults(window)

###############################################################################
# Show popup if communications not initialized
###############################################################################
if comm is None:
    sg.popup_error('Controller communications not initialized. Check INI file and hardware connection.', keep_on_top=True)

def get_comm_for_axis(axis_letter):
    """Return the appropriate comm object for the given axis."""
    if axis_letter == 'H' and comm_h is not None:
        return comm_h  # MyActuator (Servo 8)
    elif axis_letter == 'E' and comm_e is not None:
        return comm_e  # ClearCore (Servo 5)
    elif axis_letter in ['A', 'B', 'C', 'D'] and comm is not None:
        return comm  # RSI Software (Servos 1-4)
    return comm  # Fallback

def send_axis_command(axis_letter, cmd):
    """Send command to the appropriate controller based on axis."""
    controller = get_comm_for_axis(axis_letter)
    if controller is None:
        return False
    return controller.send_command(cmd)


def push_motion_defaults_to_controller(window):
    """Send saved speed/accel/decel to the controller at startup so the TIM service
    has correct values without the user having to click OK on each field."""
    for i in range(1, 5):  # Axes A-D (Servos 1-4) only
        axis_letter = AXIS_LETTERS[i - 1]
        if not hasattr(window, '_last_setpoints') or len(window._last_setpoints) < i:
            continue
        sp = window._last_setpoints[i - 1]
        scaling = AXIS_UNITS.get(axis_letter, {}).get('scaling', 1)
        gearbox = AXIS_UNITS.get(axis_letter, {}).get('gearbox', 1)
        for field, cmd_prefix in (('speed', 'SP'), ('accel', 'AC'), ('decel', 'DC')):
            val = sp.get(field)
            if val is None:
                continue
            try:
                pulses = int(round(float(val) * scaling * gearbox))
                send_axis_command(axis_letter, f'{cmd_prefix}{axis_letter}={pulses}')
            except Exception as e:
                logging.warning(f'Startup push {cmd_prefix} for axis {axis_letter} failed: {e}')


push_motion_defaults_to_controller(window)


def sync_axis_e_actual_from_commanded(window, servo_num=5):
    """Servo E has no encoder feedback; mirror the commanded endpoint into Actual Position fields."""
    try:
        if comm_e is None:
            return
        axis_letter = AXIS_LETTERS[servo_num - 1]
        if axis_letter != 'E':
            return

        commanded_pulses = getattr(comm_e, 'clearcore_commanded_position', None)
        if commanded_pulses is None:
            commanded_pulses = getattr(comm_e, 'clearcore_last_position', None)
        if commanded_pulses is None:
            return

        axis_units = AXIS_UNITS.get('E', {})
        scaling = axis_units.get('scaling', 1) or 1
        gearbox = axis_units.get('gearbox', 1) or 1
        commanded_deg = float(commanded_pulses) / (scaling * gearbox)
        commanded_disp = 0 if abs(commanded_deg) < 1e-6 else round(commanded_deg, 2)

        if not hasattr(window, '_last_valid_pos'):
            window._last_valid_pos = [''] * 8
        if not hasattr(window, '_last_pos_update_ts'):
            window._last_pos_update_ts = [None] * 8

        if f'S{servo_num}_actual_pos' in window.AllKeysDict:
            window[f'S{servo_num}_actual_pos'].update(str(commanded_disp))
        if f'S{servo_num}_actual_pos_pulses' in window.AllKeysDict:
            window[f'S{servo_num}_actual_pos_pulses'].update(str(int(round(float(commanded_pulses)))))

        window._last_valid_pos[servo_num - 1] = str(commanded_disp)
        window._last_pos_update_ts[servo_num - 1] = time.time()
        update_setpoint_highlight(window, servo_num, commanded_deg)
    except Exception:
        pass


def adjust_axis_e_actual_by_delta(window, servo_num, delta_deg):
    """Servo E helper: adjust displayed Actual Position by a known commanded delta (deg)."""
    try:
        axis_letter = AXIS_LETTERS[servo_num - 1]
        if axis_letter != 'E':
            return

        # [CHANGE 2026-03-27 11:35:00 -04:00] Use ClearCore commanded cache as baseline to avoid UI/poll jitter.
        base_deg = None
        try:
            if comm_e is not None:
                commanded_pulses = getattr(comm_e, 'clearcore_commanded_position', None)
                if commanded_pulses is None:
                    commanded_pulses = getattr(comm_e, 'clearcore_last_position', None)
                if commanded_pulses is not None:
                    axis_units = AXIS_UNITS.get('E', {})
                    scaling = axis_units.get('scaling', 1) or 1
                    gearbox = axis_units.get('gearbox', 1) or 1
                    base_deg = float(commanded_pulses) / (scaling * gearbox)
        except Exception:
            base_deg = None

        if base_deg is None:
            try:
                base_deg = float(window._last_valid_pos[servo_num - 1]) if window._last_valid_pos[servo_num - 1] not in ('', None) else None
            except Exception:
                base_deg = None
        if base_deg is None:
            # Fall back to current commanded cache if UI baseline is unavailable.
            sync_axis_e_actual_from_commanded(window, servo_num)
            try:
                base_deg = float(window._last_valid_pos[servo_num - 1]) if window._last_valid_pos[servo_num - 1] not in ('', None) else 0.0
            except Exception:
                base_deg = 0.0

        new_deg = base_deg + float(delta_deg)
        axis_units = AXIS_UNITS.get('E', {})
        scaling = axis_units.get('scaling', 1) or 1
        gearbox = axis_units.get('gearbox', 1) or 1
        new_pulses = int(round(new_deg * scaling * gearbox))

        if comm_e is not None:
            setattr(comm_e, 'clearcore_last_position', new_pulses)
            setattr(comm_e, 'clearcore_commanded_position', new_pulses)

        new_disp = 0 if abs(new_deg) < 1e-6 else round(new_deg, 2)
        if f'S{servo_num}_actual_pos' in window.AllKeysDict:
            window[f'S{servo_num}_actual_pos'].update(str(new_disp))
        if f'S{servo_num}_actual_pos_pulses' in window.AllKeysDict:
            window[f'S{servo_num}_actual_pos_pulses'].update(str(new_pulses))

        if not hasattr(window, '_last_valid_pos'):
            window._last_valid_pos = [''] * 8
        if not hasattr(window, '_last_pos_update_ts'):
            window._last_pos_update_ts = [None] * 8
        window._last_valid_pos[servo_num - 1] = str(new_disp)
        window._last_pos_update_ts[servo_num - 1] = time.time()
        update_setpoint_highlight(window, servo_num, new_deg)
    except Exception:
        pass

window_closed = False
import time
# Import the polling thread from ControllerPolling
from ControllerPolling import start_polling_thread, start_comm_health_thread

def initialize_setpoints_from_controller(window, comm):
    """Query controller for current setpoints/status and seed GUI fields."""
    if not comm:
        return
    # Ensure tracking structure exists even if controller queries fail
    if not hasattr(window, '_last_setpoints') or not window._last_setpoints or len(window._last_setpoints) != 8:
        window._last_setpoints = [{f: None for f in ['speed', 'accel', 'decel', 'abs_pos', 'rel_pos', 'jog_amount']} for _ in range(8)]
    try:
        mode = getattr(comm, 'mode', None)
    except Exception:
        mode = None
    # Currently only implemented for Galil (CommMode1) where MG operands are available
    if mode != 'CommMode1':
        return
    for idx, axis_letter in enumerate(AXIS_LETTERS):
        servo_num = idx + 1
        scaling = AXIS_UNITS.get(axis_letter, {}).get('scaling', 1) or 1
        gearbox = AXIS_UNITS.get(axis_letter, {}).get('gearbox', 1) or 1
        denom = scaling * gearbox if scaling * gearbox != 0 else 1
        queries = {
            'speed': f'MG _SP{axis_letter}',
            'accel': f'MG _AC{axis_letter}',
            'decel': f'MG _DC{axis_letter}',
            'abs_pos': f'MG _TP{axis_letter}',
            'status': f'MG _MO{axis_letter}',
        }
        results = {}
        for field, cmd in queries.items():
            try:
                resp = comm.send_command(cmd)
                if isinstance(resp, str):
                    # Take the first numeric token
                    for line in resp.splitlines():
                        line = line.strip()
                        try:
                            results[field] = float(line)
                            break
                        except ValueError:
                            continue
            except Exception:
                continue
        # Update numeric fields (convert pulses to degrees)
        for field in ['speed', 'accel', 'decel', 'abs_pos']:
            if field not in results:
                # Set default value of 0 for abs_pos if not queried
                if field == 'abs_pos':
                    window[f'S{servo_num}_{field}'].update('0')
                    window._last_setpoints[servo_num - 1][field] = 0
                continue
            raw = results[field]
            val_deg = raw / denom if field in ['speed', 'accel', 'decel', 'abs_pos'] else raw
            # Clamp to configured min/max for safety on seed
            min_val, max_val = get_limits(axis_letter, field)
            if min_val is not None and max_val is not None:
                val_deg = max(min_val, min(max_val, val_deg))
            formatted = format_display_value(val_deg)
            window[f'S{servo_num}_{field}'].update(formatted)
            # Track last setpoints for cancel restore
            window._last_setpoints[servo_num - 1][field] = val_deg
        # Relative position has no direct query; reset to 0
        window[f'S{servo_num}_rel_pos'].update('0')
        window._last_setpoints[servo_num - 1]['rel_pos'] = 0
        # Update enable/disable indicator using status (Galil _MO returns 0 when disabled)
        status_val = results.get('status')
        if status_val is not None:
            enabled = status_val == 0
            if enabled:
                window[f'S{servo_num}_status_light'].update('●', text_color='#00FF00')
                window[f'S{servo_num}_status_text'].update('Enabled', text_color='#00FF00')
            else:
                window[f'S{servo_num}_status_light'].update('●', text_color='#FFFF00')
                window[f'S{servo_num}_status_text'].update('Disabled', text_color='#FFFF00')
        if not window_closed:
            window['DEBUG_LOG'].print(f'[INIT] S{servo_num} seeded from controller: {results}')
        update_mid_speed_display(window, servo_num)


def update_setpoint_highlight(window, servo_num, actual_deg=None, tolerance=0.1):
    """Highlight abs_pos: yellow if pending, green if on target, otherwise white."""
    key = f'S{servo_num}_abs_pos'
    pending = set_pending_highlight(window, servo_num, 'abs_pos')
    if pending:
        return
    try:
        target_str = window[key].get()
        if not target_str or target_str in ('-', '.'):
            window[key].update(background_color=DEFAULT_INPUT_BG)
            return
        if actual_deg is None:
            try:
                actual_deg = float(window._last_valid_pos[servo_num - 1]) if window._last_valid_pos[servo_num - 1] else None
            except Exception:
                actual_deg = None
        if actual_deg is None:
            window[key].update(background_color=DEFAULT_INPUT_BG)
            return
        target = float(target_str)
        if abs(actual_deg - target) <= tolerance:
            window[key].update(background_color=HIGHLIGHT_INPUT_BG)
        else:
            window[key].update(background_color=DEFAULT_INPUT_BG)
    except Exception:
        window[key].update(background_color=DEFAULT_INPUT_BG)


# [CHANGE 2026-03-24 13:36:00 -04:00] Press/release jog binding intentionally disabled for safety.
bind_jog_press_release(window)

# Seed GUI setpoints/status from controller before starting polling
initialize_setpoints_from_controller(window, comm)

# [CHANGE 2026-03-24 11:19:00 -04:00] Re-apply startup speed/accel/decel defaults for any unseeded fields.
apply_startup_motion_defaults(window)


def render_datapipe_preview(window, segments):
    """Render DataPipe segments into the preview box."""
    if not segments:
        window['DP_PREVIEW'].update('')
        return
    lines = []
    lines.append('Idx  Time(ms)   Axis1(deg/pulses)  Axis2  Axis3  Axis4  Axis5')
    for idx, seg in enumerate(segments, start=1):
        line_parts = [f"{idx:02d}   {seg['time_ms']:7.1f}"]
        for axis_idx, axis_letter in enumerate(['A','B','C','D','E'], start=1):
            axis_data = seg['converted'][axis_idx - 1]
            line_parts.append(f" {axis_data['deg']:7.2f}/{axis_data['pulses']:7d}")
        lines.append('  '.join(line_parts))
    window['DP_PREVIEW'].update('\n'.join(lines))


def prepare_datapipe_segments(raw_segments):
    """Clamp degrees and convert to pulses for A-E."""
    prepared = []
    for seg in raw_segments:
        converted = []
        for axis_idx in range(1, 6):
            axis_letter = _axis_letter_for_index(axis_idx)
            converted.append(_clamp_and_convert_deg_to_pulses(axis_letter, seg['axis_deg'][axis_idx - 1]))
        prepared.append({
            'time_ms': seg['time_ms'],
            'axis_deg': seg['axis_deg'],
            'converted': converted,
        })
    return prepared


def send_datapipe_contour(comm, segments, window=None):
    """Send contour data using CD/DT (uniform time) to Galil. Uses first time_ms for DT."""
    if comm is None:
        raise RuntimeError('Controller communications not initialized.')
    if not segments:
        raise ValueError('No segments to send.')
    if getattr(comm, 'mode', None) != 'CommMode1':
        raise RuntimeError('DataPipe send only supported in CommMode1 (Galil Ethernet).')
    time_ms = segments[0]['time_ms']
    if time_ms <= 0:
        raise ValueError(f'Invalid DT (must be >0 ms); got {time_ms}. Check first row time values.')
    dt_val = max(1, int(round(time_ms)))
    cmds = []
    axis_mask = ''.join(['A','B','C','D','E'])
    cmds.append('ST')  # stop before loading
    cmds.append('CM {}'.format(axis_mask))  # enable contour mode on axes A-E
    cmds.append('DT={}'.format(dt_val))
    # Build incremental deltas in pulses
    prev_pulses = [0,0,0,0,0]
    for seg in segments:
        # If time differs, warn by using first time_ms; per-axis time not supported here
        deltas = []
        for axis_idx in range(5):
            pulses = seg['converted'][axis_idx]['pulses']
            delta = pulses - prev_pulses[axis_idx]
            deltas.append(delta)
            prev_pulses[axis_idx] = pulses
        cmds.append('CD {}'.format(','.join(str(d) for d in deltas)))
    cmds.append('CD 0,0,0,0,0=0')  # terminate contour buffer
    cmds.append('BGS')
    # Send commands sequentially to avoid overrun
    for c in cmds:
        result = comm.send_command(c)
        if window and 'DEBUG_LOG' in window.AllKeysDict:
            window['DEBUG_LOG'].print(f"[DP_CMD] {c} -> {result}")
        if result is False:
            tc1 = None
            try:
                tc1 = comm.gclib.GCommand('TC1') if hasattr(comm, 'gclib') else None
            except Exception:
                tc1 = None
            detail = f' (TC1={tc1.strip()})' if tc1 else ''
            raise RuntimeError(f'Controller rejected command: {c}{detail}')


def send_datapipe_pr(comm, segments, window=None, line_speed: float = 1.0, values=None, max_rows=None, stop_event=None):
    """Send DataPipe segments as sequential PR moves (axes A-E) for step/dir setups.

    If values are provided, SP/AC/DC are set per axis using current tab entries scaled by line_speed.
    """
    if comm is None:
        raise RuntimeError('Controller communications not initialized.')
    if not segments:
        raise ValueError('No segments to send.')
    if getattr(comm, 'mode', None) != 'CommMode1':
        raise RuntimeError('DataPipe PR send only supported in CommMode1 (Galil Ethernet).')

    cmds = []
    cmds.append('ST')  # stop before loading

    if values is not None:
        for axis_idx in range(5):
            servo_num = axis_idx + 1
            axis_letter = _axis_letter_for_index(servo_num)
            speed_key = f'S{servo_num}_speed'
            accel_key = f'S{servo_num}_accel'
            decel_key = f'S{servo_num}_decel'
            try:
                base_speed = float(values.get(speed_key, ''))
                base_accel = float(values.get(accel_key, ''))
                base_decel = float(values.get(decel_key, ''))
            except Exception:
                continue
            scaled_speed = base_speed * line_speed
            scaled_accel = base_accel * line_speed
            scaled_decel = base_decel * line_speed
            min_spd, max_spd = get_limits(axis_letter, 'speed')
            min_ac, max_ac = get_limits(axis_letter, 'accel')
            min_dc, max_dc = get_limits(axis_letter, 'decel')
            if min_spd is not None:
                scaled_speed = max(min_spd, scaled_speed)
            if max_spd is not None:
                scaled_speed = min(max_spd, scaled_speed)
            if min_ac is not None:
                scaled_accel = max(min_ac, scaled_accel)
            if max_ac is not None:
                scaled_accel = min(max_ac, scaled_accel)
            if min_dc is not None:
                scaled_decel = max(min_dc, scaled_decel)
            if max_dc is not None:
                scaled_decel = min(max_dc, scaled_decel)
            axis_units = AXIS_UNITS.get(axis_letter, {})
            scaling = axis_units.get('scaling', 1) or 1
            gearbox = axis_units.get('gearbox', 1) or 1
            pulses_speed = int(round(scaled_speed * scaling * gearbox))
            pulses_accel = int(round(scaled_accel * scaling * gearbox))
            pulses_decel = int(round(scaled_decel * scaling * gearbox))
            cmds.append(f"SP{axis_letter}={pulses_speed}")
            cmds.append(f"AC{axis_letter}={pulses_accel}")
            cmds.append(f"DC{axis_letter}={pulses_decel}")
    use_segments = segments[:max_rows] if (max_rows is not None and max_rows > 0) else segments
    # Seed deltas from current actual positions (pulses) so PR deltas are relative to where the axes are now
    prev_pulses = [0, 0, 0, 0, 0]
    try:
        for axis_idx in range(5):
            axis_letter = _axis_letter_for_index(axis_idx + 1)
            resp = comm.send_command(f"MG _RP{axis_letter}")
            if isinstance(resp, str):
                for line in resp.splitlines():
                    try:
                        prev_pulses[axis_idx] = int(float(line.strip()))
                        break
                    except Exception:
                        continue
    except Exception:
        pass

    total_steps = len(use_segments)
    for seg_idx, seg in enumerate(use_segments, start=1):
        if stop_event is not None and stop_event.is_set():
            raise RuntimeError('PR send canceled.')
        deltas = []
        for axis_idx in range(5):
            pulses = seg['converted'][axis_idx]['pulses']
            delta = pulses - prev_pulses[axis_idx]
            deltas.append(delta)
            prev_pulses[axis_idx] = pulses
        if all(d == 0 for d in deltas):
            continue  # nothing to do for this segment
        pr_cmd = 'PR {}'.format(','.join(str(d) for d in deltas))
        cmds.append(pr_cmd)
        cmds.append('BG ABCDE')
        cmds.append('AM ABCDE')
        if window and 'DP_STATUS' in window.AllKeysDict:
            try:
                window.write_event_value('DP_PR_PROGRESS', (seg_idx, total_steps))
            except Exception:
                pass

    if len(cmds) == 1:  # only ST
        raise ValueError('All DataPipe segments were zero after conversion; nothing to send.')

    for c in cmds:
        result = comm.send_command(c)
        if window and 'DEBUG_LOG' in window.AllKeysDict:
            window['DEBUG_LOG'].print(f"[DP_PR_CMD] {c} -> {result}")
        if result is False:
            tc1 = None
            try:
                tc1 = comm.gclib.GCommand('TC1') if hasattr(comm, 'gclib') else None
            except Exception:
                tc1 = None
            detail = f' (TC1={tc1.strip()})' if tc1 else ''
            raise RuntimeError(f'Controller rejected command: {c}{detail}')
        # Extra safety: wait until axes A-E report idle after AM to avoid PR while running (TC1=7)
        if isinstance(c, str) and c.strip().upper().startswith('AM'):
            try:
                axes_to_check = ['A','B','C','D','E']
                import time
                end_time = time.time() + 15.0
                poll_interval = 0.015  # 15 ms for faster PR sequence
                while time.time() < end_time:
                    if stop_event is not None and stop_event.is_set():
                        raise RuntimeError('PR send canceled.')
                    busy = False
                    for ax in axes_to_check:
                        try:
                            bg_resp = comm.send_command(f"MG _BG{ax}")
                            if isinstance(bg_resp, str):
                                for line in bg_resp.splitlines():
                                    line = line.strip()
                                    try:
                                        if float(line) != 0.0:
                                            busy = True
                                            break
                                    except Exception:
                                        continue
                            if busy:
                                break
                        except Exception:
                            pass
                    if not busy:
                        break
                    time.sleep(poll_interval)
            except Exception:
                pass



# --- Define handle_servo_event before main event loop ---
def handle_servo_event(event, values):
    # Parse servo number from event string (e.g., 'S3_enable' -> 3)
    if isinstance(event, str) and event.startswith('S') and '_' in event:
        try:
            servo_num_str, action = event[1:].split('_', 1)
            servo_num = int(servo_num_str)
            axis_letter = AXIS_LETTERS[servo_num - 1]
            map_key = f'S{servo_num}_{action}'
            print(f'[DEBUG] handle_servo_event: event={event}, servo_num={servo_num}, action={action}')
        except Exception:
            print(f'[DEBUG] handle_servo_event: failed to parse event={event}')
            return

        # Always check for setpoint OK button, regardless of map_key in COMMAND_MAP
        setpoint_fields = ['speed', 'accel', 'decel', 'abs_pos', 'rel_pos']
        if not hasattr(window, '_last_setpoints'):
            window._last_setpoints = [{f: None for f in setpoint_fields} for _ in range(8)]
        for field in setpoint_fields:
            print(f'[DEBUG] Comparing action={action} to {field}_ok')
            if action == f'{field}_ok':
                print(f'[DEBUG] handle_servo_event called for {field}_ok, S{servo_num}')
                original_text = window[f'S{servo_num}_{field}'].get()
                previous_setpoint = window._last_setpoints[servo_num - 1].get(field)
                value = values.get(f'S{servo_num}_{field}', None)
                if value is None or value == '':
                    print(f'[DEBUG] No value entered for {field} (S{servo_num})')
                    sg.popup_error(f'Please enter a value for {field}', keep_on_top=True)
                    print('[DEBUG] RETURN: No value entered')
                    return
                try:
                    # Accept float input and keep one decimal place for display
                    value = round(float(value), 1)
                    formatted_value = format_display_value(value)
                    window[f'S{servo_num}_{field}'].update(formatted_value)
                except ValueError:
                    print(f'[DEBUG] Invalid value for {field} (S{servo_num}): {value}')
                    sg.popup_error(f'Invalid value for {field}', keep_on_top=True)
                    print('[DEBUG] RETURN: Invalid value')
                    return
                axis_letter = AXIS_LETTERS[servo_num - 1]
                min_val, max_val = get_limits(axis_letter, field)
                # For relative moves, ensure resulting position stays within limits
                if field == 'rel_pos':
                    try:
                        current_pos = float(window._last_valid_pos[servo_num - 1]) if window._last_valid_pos[servo_num - 1] else 0.0
                    except Exception:
                        current_pos = 0.0
                    # [CHANGE 2026-03-24 12:42:00 -04:00] Axis E safety: enforce conservative step cap without hard-blocking transient unknown live position.
                    if axis_letter == 'E':
                        if abs(value) > AXIS_E_MAX_RELATIVE_STEP_DEG:
                            sg.popup_error(
                                f'Axis E relative move blocked for safety.\n\n'
                                f'Max single relative step is ±{AXIS_E_MAX_RELATIVE_STEP_DEG} deg.\n'
                                f'Entered: {value} deg',
                                keep_on_top=True,
                                title='Safety Block'
                            )
                            return
                    target_pos = current_pos + value
                    if target_pos < min_val or target_pos > max_val:
                        print(f'[DEBUG] Relative move would exceed limits: current={current_pos}, delta={value}, target={target_pos}, limits=({min_val},{max_val})')
                        sg.popup_error(f"Move would exceed limits ({min_val} to {max_val}). Current: {current_pos}, Target: {target_pos}", keep_on_top=True)
                        return
                if value < min_val or value > max_val:
                    print(f'[DEBUG] Value for {field} (S{servo_num}) out of range: {value}')
                    sg.popup_error(f"Value for {field} must be between {min_val} and {max_val}", keep_on_top=True)
                    print('[DEBUG] RETURN: Value out of range')
                    return
                # Confirm with user before sending command (optional per tab)
                confirm_required = bool(values.get(f'S{servo_num}_confirm_ok', True))
                confirm_label = field.replace('_', ' ').title()
                confirm = 'OK'
                if confirm_required:
                    confirm = sg.popup_ok_cancel(
                        f"Send setpoint for {confirm_label} (S{servo_num})?\nValue: {formatted_value}",
                        keep_on_top=True,
                        title='Confirm Setpoint'
                    )
                if confirm != 'OK':
                    restore_val = previous_setpoint if previous_setpoint is not None else original_text
                    window[f'S{servo_num}_{field}'].update(format_display_value(restore_val) if restore_val not in (None, '') else '')
                    if not window_closed:
                        window['DEBUG_LOG'].print(f"[INFO] Setpoint canceled for {confirm_label} S{servo_num}; reverted to {format_display_value(restore_val) if restore_val not in (None, '') else 'blank'}\n", end='')
                        window['DEBUG_LOG'].Widget.see('end')
                        if field == 'abs_pos':
                            update_setpoint_highlight(window, servo_num)
                        else:
                            set_pending_highlight(window, servo_num, field)
                    print('[DEBUG] User canceled setpoint send')
                    return
                # Convert engineering units to pulses using axis scaling and gearbox.
                scaling = AXIS_UNITS[axis_letter].get('scaling', 1)
                gearbox = AXIS_UNITS[axis_letter].get('gearbox', 1)
                pulses_value = int(round(value * scaling * gearbox))

                cmd_func = COMMAND_MAP.get(f'S{servo_num}_{field}')
                if callable(cmd_func):
                    cmd = cmd_func(pulses_value)
                else:
                    cmd = cmd_func

                # ClearCore Axis E: stage move on OK, execute on Start Motion
                if axis_letter == 'E' and field in ('abs_pos', 'rel_pos') and isinstance(cmd, str):
                    if cmd.startswith('PAE='):
                        cmd = 'QPAE=' + cmd.split('=', 1)[1]
                    elif cmd.startswith('PRE='):
                        cmd = 'QPRE=' + cmd.split('=', 1)[1]

                print(f'[DEBUG] About to send setpoint command: {cmd}')
                if not cmd:
                    print('[DEBUG] RETURN: cmd is None')
                    return
                controller = get_comm_for_axis(axis_letter)
                if not controller:
                    print('[DEBUG] RETURN: controller is None')
                    sg.popup_error('Controller communications not initialized.', keep_on_top=True)
                    return
                try:
                    response = send_axis_command(axis_letter, cmd)
                    print(f'[DEBUG] Setpoint command sent, response: {response}')
                    if not window_closed:
                        if axis_letter == 'E' and field in ('abs_pos', 'rel_pos') and isinstance(cmd, str) and cmd.startswith('QP'):
                            log_line = (
                                f"[TEST LOG] {field.capitalize()} OK for S{servo_num}: Staged {cmd} "
                                f"(waiting for Start Motion)\nReply: {response}\n"
                            )
                        else:
                            log_line = f"[TEST LOG] {field.capitalize()} OK for S{servo_num}: Sent {cmd}\nReply: {response}\n"
                        print(f'[DEBUG] Logging setpoint to DEBUG_LOG: {log_line.strip()}')
                        window['DEBUG_LOG'].print(log_line, end='')
                        window['DEBUG_LOG'].Widget.see('end')
                        window.refresh()
                    # Persist last confirmed setpoint value for cancel restores
                    window._last_setpoints[servo_num - 1][field] = value
                    # Track command type for Start Motion safety
                    if field == 'abs_pos':
                        LAST_MOTION_COMMAND[servo_num - 1] = 'abs'
                        update_setpoint_highlight(window, servo_num)
                    elif field == 'rel_pos':
                        LAST_MOTION_COMMAND[servo_num - 1] = 'rel'
                        set_pending_highlight(window, servo_num, field)
                    else:
                        set_pending_highlight(window, servo_num, field)
                    if field in ('speed', 'accel', 'decel'):
                        # Persist latest motion tuning immediately after a successful update.
                        try:
                            save_motion_defaults_from_values(values)
                        except Exception:
                            pass
                        update_mid_speed_display(window, servo_num)
                except Exception as e:
                    import traceback
                    error_details = f'{e}\n' + traceback.format_exc()
                    print(f'[DEBUG] Exception sending setpoint command: {error_details}')
                    sg.popup_error(f'Error sending command: {e}', keep_on_top=True)
                    if not window_closed:
                        window['DEBUG_LOG'].update(f'[ERROR] Error sending command: {error_details}\n', append=True)
                        window['DEBUG_LOG'].Widget.see('end')
                return
        # Only handle direct motor control buttons if not a setpoint OK event
        if map_key in COMMAND_MAP:
            cmd = None
            match action:
                case 'jog':
                    speed_val = values.get(f'S{servo_num}_speed', None)
                    if speed_val is None or speed_val == '':
                        sg.popup_error('Please enter a speed value for Jog', keep_on_top=True)
                        return
                    try:
                        speed_val = float(speed_val)
                    except ValueError:
                        sg.popup_error('Invalid speed value for Jog', keep_on_top=True)
                        return
                    # Soft limit: prevent jogging past min/max if current position known
                    axis_letter = AXIS_LETTERS[servo_num - 1]
                    try:
                        current_pos = float(window._last_valid_pos[servo_num - 1]) if window._last_valid_pos[servo_num - 1] else None
                    except Exception:
                        current_pos = None
                    min_val, max_val = AXIS_UNITS[axis_letter]['min'], AXIS_UNITS[axis_letter]['max']
                    if current_pos is not None:
                        if speed_val > 0 and current_pos >= max_val:
                            if hasattr(window, '_jog_limit_hit'):
                                window._jog_limit_hit[servo_num - 1] = True
                            window[f'S{servo_num}_status_light'].update('●', text_color='#FFA500')
                            window[f'S{servo_num}_status_text'].update('At Max Limit', text_color='#FFA500')
                            sg.popup_error(f'Jog blocked: at limit {max_val} deg', keep_on_top=True)
                            return
                        if speed_val < 0 and current_pos <= min_val:
                            if hasattr(window, '_jog_limit_hit'):
                                window._jog_limit_hit[servo_num - 1] = True
                            window[f'S{servo_num}_status_light'].update('●', text_color='#FFA500')
                            window[f'S{servo_num}_status_text'].update('At Min Limit', text_color='#FFA500')
                            sg.popup_error(f'Jog blocked: at limit {min_val} deg', keep_on_top=True)
                            return
                    # Convert degrees/sec to pulses/sec using scaling and gearbox.
                    axis_letter = AXIS_LETTERS[servo_num - 1]
                    scaling = AXIS_UNITS[axis_letter].get('scaling', 1)
                    gearbox = AXIS_UNITS[axis_letter].get('gearbox', 1)
                    speed_val = int(round(speed_val * scaling * gearbox))
                    cmd_func = COMMAND_MAP[map_key]
                    if callable(cmd_func):
                        cmd = cmd_func(speed_val)
                    else:
                        cmd = cmd_func
                    LAST_MOTION_COMMAND[servo_num - 1] = 'jog'
                case 'start':
                    # [CHANGE 2026-04-17 00:00:00 -04:00] Added diagnostic print so start execution is visible in terminal log.
                    print(f'[DEBUG] Start case entered for S{servo_num}, axis={axis_letter}, LAST_MOTION_COMMAND={LAST_MOTION_COMMAND[servo_num - 1]}')
                    # SAFETY: Block Start Motion if no valid position command was set
                    last_cmd = LAST_MOTION_COMMAND[servo_num - 1]
                    if last_cmd not in ('abs', 'rel'):
                        sg.popup_error(
                            f'Start Motion blocked for safety.\n\n'
                            f'You must set an Absolute or Relative position\n'
                            f'before using Start Motion.\n\n'
                            f'Current state: {last_cmd or "no position set"}',
                            keep_on_top=True,
                            title='Safety Block'
                        )
                        return
                    if axis_letter == 'E':
                        try:
                            axis_units = AXIS_UNITS[axis_letter]
                            scaling = axis_units.get('scaling', 1) or 1
                            gearbox = axis_units.get('gearbox', 1) or 1

                            # [CHANGE 2026-03-24 11:06:00 -04:00] Deterministically stage Axis E target on Start.
                            # This prevents BGE from running without a pending target when operator hasn't pressed setpoint OK recently.
                            if last_cmd == 'abs':
                                abs_raw = values.get(f'S{servo_num}_abs_pos', '')
                                if abs_raw not in ('', None, '-', '.'):
                                    abs_deg = float(abs_raw)
                                    abs_min, abs_max = get_limits(axis_letter, 'abs_pos')
                                    abs_deg = max(abs_min, min(abs_max, abs_deg))
                                    abs_pulses = int(round(abs_deg * scaling * gearbox))
                                    send_axis_command(axis_letter, f'QPAE={abs_pulses}')
                                    if comm_e is not None:
                                        setattr(comm_e, 'clearcore_commanded_position', abs_pulses)
                            elif last_cmd == 'rel':
                                rel_raw = values.get(f'S{servo_num}_rel_pos', '')
                                if rel_raw not in ('', None, '-', '.'):
                                    rel_deg = float(rel_raw)
                                    rel_pulses = int(round(rel_deg * scaling * gearbox))
                                    send_axis_command(axis_letter, f'QPRE={rel_pulses}')
                                    if comm_e is not None:
                                        base = getattr(comm_e, 'clearcore_commanded_position', None)
                                        if base is None:
                                            base = getattr(comm_e, 'clearcore_last_position', 0)
                                        setattr(comm_e, 'clearcore_commanded_position', int(base) + rel_pulses)

                            speed_raw = values.get(f'S{servo_num}_speed', '')
                            accel_raw = values.get(f'S{servo_num}_accel', '')

                            if speed_raw not in ('', None, '-', '.'):
                                speed_val = float(speed_raw)
                                speed_pulses = int(round(speed_val * scaling * gearbox))
                                send_axis_command(axis_letter, f'SP{axis_letter}={speed_pulses}')

                            if accel_raw not in ('', None, '-', '.'):
                                accel_val = float(accel_raw)
                                accel_pulses = int(round(accel_val * scaling * gearbox))
                                send_axis_command(axis_letter, f'AC{axis_letter}={accel_pulses}')
                        except Exception as start_param_err:
                            if not window_closed:
                                window['DEBUG_LOG'].print(f'[WARN] Axis E start pre-load skipped: {start_param_err}')
                    cmd = COMMAND_MAP[map_key] if not callable(COMMAND_MAP[map_key]) else None
                    print(f'[DEBUG] Start: cmd={cmd!r} for S{servo_num} (axis {axis_letter})')
                case 'enable' | 'disable' | 'stop':
                    cmd = COMMAND_MAP[map_key] if not callable(COMMAND_MAP[map_key]) else None
                    if action == 'stop':
                        # Clear motion command tracking on stop
                        LAST_MOTION_COMMAND[servo_num - 1] = None
                        # [CHANGE 2026-03-24 10:58:00 -04:00] Force-cancel any active press-and-hold jog worker before issuing stop.
                        idx = servo_num - 1
                        if 0 <= idx < len(JOG_STOP_EVENTS):
                            try:
                                stop_evt = JOG_STOP_EVENTS[idx]
                                if stop_evt is not None:
                                    stop_evt.set()
                            except Exception:
                                pass
                            JOG_STOP_EVENTS[idx] = None
                case _:
                    # For any other actions, fallback to original logic if needed
                    cmd = COMMAND_MAP[map_key] if not callable(COMMAND_MAP[map_key]) else None
            if cmd:
                controller = get_comm_for_axis(axis_letter)
                if controller:
                    try:
                        # If disabling, send stop command first
                        if action == 'disable':
                            stop_key = f'S{servo_num}_stop'
                            stop_cmd = COMMAND_MAP.get(stop_key)
                            if stop_cmd and axis_letter != 'E':
                                stop_cmd_val = stop_cmd if not callable(stop_cmd) else stop_cmd()
                                send_axis_command(axis_letter, stop_cmd_val)
                        response = send_axis_command(axis_letter, cmd)
                        # Log request and reply in DEBUG_LOG for all actions
                        if not window_closed:
                            prev_log = window['DEBUG_LOG'].get()
                            new_log = f"[TEST LOG] {action.capitalize()} button clicked for S{servo_num}: Sent {cmd}\nReply: {response}\n"
                            window['DEBUG_LOG'].update(prev_log + new_log)
                        action_succeeded = not (
                            response is False or
                            (isinstance(response, str) and str(response).strip().upper().startswith('UNSUPPORTED'))
                        )
                        if action_succeeded:
                            # Immediately update indicator color (bright green for enable, bright yellow for disable)
                            match action:
                                case 'enable':
                                    window[f'S{servo_num}_status_light'].update('●', text_color='#00FF00')  # Bright green
                                    window[f'S{servo_num}_status_text'].update('Enabled', text_color='#00FF00')
                                case 'disable':
                                    window[f'S{servo_num}_status_light'].update('●', text_color='#FFFF00')  # Bright yellow
                                    window[f'S{servo_num}_status_text'].update('Disabled', text_color='#FFFF00')
                            if axis_letter == 'E' and action == 'start':
                                # [CHANGE 2026-03-27 11:05:00 -04:00] Servo E has no feedback; mirror accepted start target.
                                sync_axis_e_actual_from_commanded(window, servo_num)
                        elif action in ('disable', 'stop') and axis_letter == 'E':
                            # [CHANGE 2026-03-23 16:32:24 -04:00] Non-blocking unsupported indicator for Axis E disable/stop.
                            if not window_closed:
                                window['DEBUG_LOG'].print(f"[WARN] Axis E {action} is not supported by current ClearCore firmware command set.")
                            window[f'S{servo_num}_status_light'].update('●', text_color='#FFA500')
                            window[f'S{servo_num}_status_text'].update(f'{action.capitalize()} unsupported', text_color='#FFA500')
                    except Exception as e:
                        import traceback
                        error_details = f'{e}\n' + traceback.format_exc()
                        sg.popup_error(f'Error sending command: {e}', keep_on_top=True)
                        if not window_closed:
                            prev_log = window['DEBUG_LOG'].get()
                            window['DEBUG_LOG'].update(prev_log + f'[ERROR] Error sending command: {error_details}\n')
                else:
                    sg.popup_error('Controller communications not initialized.', keep_on_top=True)
            return


def handle_all_tab_event(window, event, values):
    """Legacy placeholder: per-row OKs removed."""
    return False


def wait_for_axis_complete(comm, axis_letter, timeout=15.0, poll_interval=0.1, stop_event=None):
    """Poll _BG<axis> until motion complete or timeout; stop_event is ignored for graceful finishes."""
    import time
    end_time = time.time() + timeout
    while time.time() < end_time:
        try:
            resp = comm.send_command(f'MG _BG{axis_letter}')
            if resp is None:
                time.sleep(poll_interval)
                continue
            try:
                val = float(str(resp).strip())
            except Exception:
                val = 1.0
            if val == 0:
                return True
        except Exception:
            pass
        time.sleep(poll_interval)
    return False


def handle_all_run_sequence(window, comm, values):
    """Start sequence run in background; supports optional repeat until stopped."""
    global SEQ_THREAD, SEQ_STOP_EVENT, SEQ_RUNNING

    if comm is None:
        sg.popup_error('Controller communications not initialized.', keep_on_top=True)
        return
    if SEQ_RUNNING:
        sg.popup_error('Sequence already running. Stop it before starting again.', keep_on_top=True)
        return

    selected = []
    for i, axis_letter in enumerate(AXIS_LETTERS, start=1):
        if not values.get(f'ALL_S{i}_enabled', False):
            continue
        setpoints = []
        for pos_field in ['pos1', 'pos2', 'pos3', 'pos4', 'pos5']:
            key = f'ALL_S{i}_{pos_field}'
            raw = str(values.get(key, '')).strip()
            if raw in ('', '-', '.', '-.'):
                setpoints.append(None)
                continue
            try:
                setpoints.append(float(raw))
            except ValueError:
                sg.popup_error(f'Invalid setpoint for Servo {i} {pos_field}: {raw}', keep_on_top=True)
                return
        if any(sp is not None for sp in setpoints):
            axis_units = AXIS_UNITS.get(axis_letter, {})
            selected.append({
                'servo_num': i,
                'axis_letter': axis_letter,
                'setpoints': setpoints,
                'min': axis_units.get('min', NUMERIC_LIMITS['abs_pos'][0]),
                'max': axis_units.get('max', NUMERIC_LIMITS['abs_pos'][1]),
                'scaling': axis_units.get('scaling', 1) or 1,
                'gearbox': axis_units.get('gearbox', 1) or 1,
            })

    if not selected:
        sg.popup_error('No enabled servos with setpoints to run.', keep_on_top=True)
        return

    for entry in selected:
        for idx, sp in enumerate(entry['setpoints'], start=1):
            if sp is None:
                continue
            if sp < entry['min'] or sp > entry['max']:
                sg.popup_error(f"Servo {entry['servo_num']} setpoint {idx} out of range ({entry['min']} to {entry['max']}).", keep_on_top=True)
                return

    # Validate line speed (0-2)
    try:
        line_speed = float(str(values.get('ALL_LINE_SPEED', '1')).strip() or '1')
    except Exception:
        sg.popup_error('Invalid Line Speed. Enter a number between 0 and 2.', keep_on_top=True)
        return
    if line_speed < 0 or line_speed > 2:
        sg.popup_error('Line Speed must be between 0 and 2.', keep_on_top=True)
        return

    repeat_flag = bool(values.get('ALL_REPEAT', False))
    # Persist sequence inputs for next launch
    save_sequence_state_from_values(values)
    # Reset displayed step times
    for idx in range(1, 6):
        key = f'ALL_STEP{idx}_TIME'
        if key in window.AllKeysDict:
            window[key].update('—')
    SEQ_STOP_EVENT = threading.Event()
    SEQ_RUNNING = True
    if 'ALL_STOP_SEQUENCE' in window.AllKeysDict:
        window['ALL_STOP_SEQUENCE'].update(disabled=False)
    if 'ALL_RUN_SEQUENCE' in window.AllKeysDict:
        window['ALL_RUN_SEQUENCE'].update(disabled=True)

    def seq_log(msg):
        try:
            window.write_event_value('ALL_SEQ_LOG', msg)
        except Exception:
            pass

    def run_once():
        import time
        if SEQ_STOP_EVENT is not None and SEQ_STOP_EVENT.is_set():
            return False
        for idx in range(len(selected[0]['setpoints'])):  # all setpoints (now 4)
            if SEQ_STOP_EVENT is not None and SEQ_STOP_EVENT.is_set():
                return False
            step_start = time.perf_counter()
            sent_axes = []
            for entry in selected:
                sp = entry['setpoints'][idx]
                if sp is None:
                    continue
                # Fetch and scale speed using Line Speed multiplier
                speed_key = f"S{entry['servo_num']}_speed"
                speed_str = values.get(speed_key, '')
                if speed_str in ('', '-', '.'):  # require a speed
                    window.write_event_value('ALL_SEQ_ERROR', f"Missing speed for Servo {entry['servo_num']} (tab speed field).")
                    return False
                try:
                    base_speed = float(speed_str)
                except Exception:
                    window.write_event_value('ALL_SEQ_ERROR', f"Invalid speed for Servo {entry['servo_num']}: {speed_str}")
                    return False
                scaled_speed = base_speed * line_speed
                min_spd, max_spd = get_limits(entry['axis_letter'], 'speed')
                if min_spd is not None:
                    scaled_speed = max(min_spd, scaled_speed)
                if max_spd is not None:
                    scaled_speed = min(max_spd, scaled_speed)

                # Fetch and scale accel/decel using Line Speed multiplier
                accel_key = f"S{entry['servo_num']}_accel"
                decel_key = f"S{entry['servo_num']}_decel"
                accel_str = values.get(accel_key, '')
                decel_str = values.get(decel_key, '')
                try:
                    base_accel = float(accel_str)
                    base_decel = float(decel_str)
                except Exception:
                    window.write_event_value('ALL_SEQ_ERROR', f"Invalid accel/decel for Servo {entry['servo_num']} (check tab fields).")
                    return False
                scaled_accel = base_accel * line_speed
                scaled_decel = base_decel * line_speed
                min_ac, max_ac = get_limits(entry['axis_letter'], 'accel')
                min_dc, max_dc = get_limits(entry['axis_letter'], 'decel')
                if min_ac is not None:
                    scaled_accel = max(min_ac, scaled_accel)
                if max_ac is not None:
                    scaled_accel = min(max_ac, scaled_accel)
                if min_dc is not None:
                    scaled_decel = max(min_dc, scaled_decel)
                if max_dc is not None:
                    scaled_decel = min(max_dc, scaled_decel)

                pulses_speed = int(round(scaled_speed * entry['scaling'] * entry['gearbox']))
                pulses_accel = int(round(scaled_accel * entry['scaling'] * entry['gearbox']))
                pulses_decel = int(round(scaled_decel * entry['scaling'] * entry['gearbox']))
                # Send speed/accel/decel before position
                try:
                    comm.send_command(f"SP{entry['axis_letter']}={pulses_speed}")
                    comm.send_command(f"AC{entry['axis_letter']}={pulses_accel}")
                    comm.send_command(f"DC{entry['axis_letter']}={pulses_decel}")
                except Exception as ex:
                    window.write_event_value('ALL_SEQ_ERROR', f"Error sending speed/accel/decel for Servo {entry['servo_num']}: {ex}")
                    return False
                pulses_value = int(round(sp * entry['scaling'] * entry['gearbox']))
                cmd = f"PA{entry['axis_letter']}={pulses_value};BG{entry['axis_letter']}"
                try:
                    resp = comm.send_command(cmd)
                    sent_axes.append((entry['servo_num'], entry['axis_letter']))
                    seq_log(f"[SEQ] S{entry['servo_num']} Setpoint {idx+1}: Sent {cmd} -> {resp}")
                except Exception as ex:
                    window.write_event_value('ALL_SEQ_ERROR', f"Error sending setpoint {idx+1} for Servo {entry['servo_num']}: {ex}")
                    return False
            for servo_num, axis_letter in sent_axes:
                if SEQ_STOP_EVENT is not None and SEQ_STOP_EVENT.is_set():
                    return False
                if not wait_for_axis_complete(comm, axis_letter):
                    window.write_event_value('ALL_SEQ_ERROR', f'Servo {servo_num} did not complete setpoint {idx+1} in time.')
                    return False
            step_elapsed = time.perf_counter() - step_start
            try:
                window.write_event_value('ALL_STEP_TIME', (idx + 1, step_elapsed))
            except Exception:
                pass
        return True

    def worker():
        status = 'completed'
        try:
            while True:
                if SEQ_STOP_EVENT is not None and SEQ_STOP_EVENT.is_set():
                    status = 'stopped'
                    break
                ok = run_once()
                if not ok:
                    status = 'stopped' if SEQ_STOP_EVENT is not None and SEQ_STOP_EVENT.is_set() else 'failed'
                    break
                if not repeat_flag or (SEQ_STOP_EVENT is not None and SEQ_STOP_EVENT.is_set()):
                    break
            if SEQ_STOP_EVENT is not None and SEQ_STOP_EVENT.is_set() and status == 'completed' and repeat_flag:
                status = 'stopped'
        except Exception as ex:
            status = f'error: {ex}'
            window.write_event_value('ALL_SEQ_ERROR', f'Sequence error: {ex}')
        finally:
            window.write_event_value('ALL_SEQ_DONE', status)

    SEQ_THREAD = threading.Thread(target=worker, daemon=True)
    SEQ_THREAD.start()


def handle_jog_press(window, servo_num, direction, is_press, values):
    """Start jog on press and stop on release for Jog CW/CCW buttons."""
    try:
        axis_letter = AXIS_LETTERS[servo_num - 1]
    except Exception:
        print(f'[DEBUG] Invalid servo_num for jog: {servo_num}')
        return
    
    # Get the correct comm object for this axis
    comm = get_comm_for_axis(axis_letter)
    if comm is None:
        print(f'[DEBUG] No comm object for axis {axis_letter}; cannot jog')
        return

    speed_str = values.get(f'S{servo_num}_speed', '')
    if speed_str in ('', '-', '.'):
        print('[DEBUG] No speed set; ignoring jog press')
        return
    try:
        speed_val = float(speed_str)
    except ValueError:
        print(f'[DEBUG] Invalid speed value for Jog: {speed_str}')
        return

    jog_amount_str = str(values.get(f'S{servo_num}_jog_amount', JOG_STEP_DEFAULT_DEG)).strip()
    if jog_amount_str in ('', '-', '.', '-.'):
        jog_amount_deg = JOG_STEP_DEFAULT_DEG
    else:
        try:
            jog_amount_deg = float(jog_amount_str)
        except ValueError:
            sg.popup_error(
                f'Jog amount must be numeric ({JOG_STEP_MIN_DEG} to {JOG_STEP_MAX_DEG} deg).',
                keep_on_top=True
            )
            return

    if jog_amount_deg < JOG_STEP_MIN_DEG or jog_amount_deg > JOG_STEP_MAX_DEG:
        sg.popup_error(
            f'Jog amount must be between {JOG_STEP_MIN_DEG} and {JOG_STEP_MAX_DEG} deg.',
            keep_on_top=True
        )
        try:
            window[f'S{servo_num}_jog_amount'].update(format_display_value(JOG_STEP_DEFAULT_DEG))
        except Exception:
            pass
        return

    sign = 1 if str(direction).lower() == 'cw' else -1
    axis_units = AXIS_UNITS.get(axis_letter, {})
    if axis_units.get('reverse', False):
        sign = -sign
    signed_speed = speed_val * sign
    scaling = axis_units.get('scaling', 1) or 1
    gearbox = axis_units.get('gearbox', 1) or 1
    min_val = axis_units.get('min', NUMERIC_LIMITS['abs_pos'][0])
    max_val = axis_units.get('max', NUMERIC_LIMITS['abs_pos'][1])

    try:
        current_pos = float(window._last_valid_pos[servo_num - 1]) if window._last_valid_pos[servo_num - 1] else None
    except Exception:
        current_pos = None

    pos_is_fresh = True
    try:
        if hasattr(window, '_last_pos_update_ts') and len(window._last_pos_update_ts) >= servo_num:
            ts = window._last_pos_update_ts[servo_num - 1]
            pos_is_fresh = (ts is not None) and ((time.time() - float(ts)) <= 1.5)
    except Exception:
        pos_is_fresh = True

    # Don't enforce pre-jog limits on stale positions — dirty startup actuals
    # can falsely show 0° and block valid jogs. POSITION_POLL enforces limits at runtime.
    enforce_limit = current_pos is not None and axis_letter != 'E' and pos_is_fresh

    if enforce_limit:
        if signed_speed > 0 and current_pos >= max_val:
            if hasattr(window, '_jog_limit_hit'):
                window._jog_limit_hit[servo_num - 1] = True
            window[f'S{servo_num}_status_light'].update('●', text_color='#FFA500')
            window[f'S{servo_num}_status_text'].update('At Max Limit', text_color='#FFA500')
            sg.popup_error(f'Jog blocked: at upper limit {max_val} deg', keep_on_top=True)
            return
        if signed_speed < 0 and current_pos <= min_val:
            if hasattr(window, '_jog_limit_hit'):
                window._jog_limit_hit[servo_num - 1] = True
            window[f'S{servo_num}_status_light'].update('●', text_color='#FFA500')
            window[f'S{servo_num}_status_text'].update('At Min Limit', text_color='#FFA500')
            sg.popup_error(f'Jog blocked: at lower limit {min_val} deg', keep_on_top=True)
            return

    # [CHANGE 2026-03-27 10:15:00 -04:00] Use GUI jog amount (deg) for one-shot jog step.
    step_pulses = int(round(jog_amount_deg * scaling * gearbox))
    if step_pulses <= 0:
        print(f'[DEBUG] Jog amount <= 0 for S{servo_num}; no move issued')
        return
    step_pulses *= sign

    if is_press:
        cmd = f'PR{axis_letter}={step_pulses}'
        try:
            response = comm.send_command(cmd)
            print(f'[DEBUG] JOG one-shot command: {cmd} -> {response}')
            # Successful jog — clear any jog limit indicator for this axis.
            if hasattr(window, '_jog_limit_hit'):
                window._jog_limit_hit[servo_num - 1] = False
                window[f'S{servo_num}_status_light'].update('●', text_color='#00FF00')
                window[f'S{servo_num}_status_text'].update('Enabled', text_color='#00FF00')

            if axis_letter in ('A', 'B', 'C', 'D'):
                # RapidCode axes: stage speed/accel/decel then fire with BG.
                # PR only stages the distance; motion profile must be explicit.
                sp_pulses = int(round(speed_val * scaling * gearbox))
                accel_str = str(values.get(f'S{servo_num}_accel', '')).strip()
                decel_str = str(values.get(f'S{servo_num}_decel', '')).strip()
                try:
                    accel_pulses = int(round(float(accel_str) * scaling * gearbox)) if accel_str else sp_pulses * 2
                    decel_pulses = int(round(float(decel_str) * scaling * gearbox)) if decel_str else sp_pulses * 2
                except ValueError:
                    accel_pulses = sp_pulses * 2
                    decel_pulses = sp_pulses * 2
                comm.send_command(f'SP{axis_letter}={sp_pulses}')
                comm.send_command(f'AC{axis_letter}={accel_pulses}')
                comm.send_command(f'DC{axis_letter}={decel_pulses}')
                bg_response = comm.send_command(f'BG{axis_letter}')
                print(f'[DEBUG] JOG BG{axis_letter} -> {bg_response}')
            elif axis_letter == 'E':
                # [CHANGE 2026-03-27 11:50:00 -04:00] Servo E PRE updates commanded cache; mirror it without applying a second delta.
                jog_ok = not (
                    response is False or
                    (isinstance(response, str) and str(response).strip().upper().startswith('UNSUPPORTED'))
                )
                if jog_ok:
                    sync_axis_e_actual_from_commanded(window, servo_num)
        except Exception as ex:
            print(f'[DEBUG] Jog one-shot failed: {ex}')
    else:
        cmd = f'ST{axis_letter}'
        try:
            response = comm.send_command(cmd)
            print(f'[DEBUG] JOG release command: {cmd} -> {response}')
        except Exception as ex:
            print(f'[DEBUG] Jog release failed: {ex}')
    return

# Start background polling threads using ControllerPolling.
# - POSITION_POLL thread updates live motion/status fields per active axis routing.
# - COMM_HEALTH thread updates per-axis link indicators (Comms OK / No Link).
# Both threads post thread-safe events to the GUI event loop.

# [CHANGE 2026-03-24 14:54:00 -04:00] Include comm_e so Axis E polling uses ClearCore path.
polling_thread, polling_stop_event = start_polling_thread(window, comm, comm_e, comm_h)

# [CHANGE 2026-04-17 00:00:00 -04:00] Start comm health thread: pings each controller every 5s and updates link indicators.
comm_health_thread, comm_health_stop_event = start_comm_health_thread(window, comm, comm_e, comm_h, interval=5.0)

# Main event loop (no periodic polling here)
while True:
    try:
        event, values = window.read(timeout=100)
    except Exception as loop_error:
        try:
            sg.popup_error(f'UI loop error: {loop_error}', keep_on_top=True)
        except Exception:
            pass
        continue
    if event == 'SHOW_POLL_LOGS':
        LOG_POSITION_POLLS = bool(values.get('SHOW_POLL_LOGS', False))
        continue
    if event == 'TABGROUP':
        _refresh_description_colors(window)
        continue
    if event == 'ESTOP':
        # Immediate stop for all axes
        # [CHANGE 2026-03-24 16:18:00 -04:00] Send explicit per-axis stops for mixed-controller axes (E/H) in addition to global ST.
        # Cancel any in-flight DataPipe PR send
        if hasattr(window, '_dp_pr_stop') and window._dp_pr_stop:
            try:
                window._dp_pr_stop.set()
            except Exception:
                pass
        if SEQ_STOP_EVENT is not None:
            try:
                SEQ_STOP_EVENT.set()
            except Exception:
                pass
        SEQ_RUNNING = False
        if 'ALL_RUN_SEQUENCE' in window.AllKeysDict:
            window['ALL_RUN_SEQUENCE'].update(disabled=False)
        if 'ALL_STOP_SEQUENCE' in window.AllKeysDict:
            window['ALL_STOP_SEQUENCE'].update(disabled=True)
        stop_errors = []
        # Main RSI/Galil path (A-D and any axes mapped there)
        if comm:
            try:
                resp = comm.send_command('ST')
                if not window_closed:
                    window['DEBUG_LOG'].print(f'[ESTOP] Sent ST to main controller -> {resp}')
            except Exception as ex:
                stop_errors.append(f'main ST failed: {ex}')

        # Explicit per-axis stop for mixed-controller axes
        for axis_letter, servo_num in [('E', 5), ('H', 8)]:
            try:
                stop_key = f'S{servo_num}_stop'
                stop_cmd = COMMAND_MAP.get(stop_key)
                if stop_cmd:
                    stop_cmd_val = stop_cmd if not callable(stop_cmd) else stop_cmd()
                    stop_resp = send_axis_command(axis_letter, stop_cmd_val)
                    if not window_closed:
                        window['DEBUG_LOG'].print(f'[ESTOP] Sent {stop_cmd_val} to Axis {axis_letter} -> {stop_resp}')
            except Exception as ex:
                stop_errors.append(f'Axis {axis_letter} stop failed: {ex}')

        LAST_MOTION_COMMAND[:] = [None]*8
        if not window_closed:
            for idx in range(1, 9):
                window[f'S{idx}_status_light'].update('●', text_color='#FF0000')
                window[f'S{idx}_status_text'].update('E-STOP', text_color='#FF0000')

        if stop_errors and not window_closed:
            sg.popup_error('E-STOP completed with errors:\n' + '\n'.join(stop_errors), keep_on_top=True)
        elif (comm is None and comm_e is None and comm_h is None) and not window_closed:
            sg.popup_error('Controller communications not initialized.', keep_on_top=True)
        continue
    if event == 'ALL_SEQ_LOG':
        msg = values.get(event, '')
        if not window_closed:
            window['DEBUG_LOG'].print(msg)
        continue
    if event == 'ALL_SEQ_ERROR':
        err_msg = values.get(event, '')
        if not window_closed:
            window['DEBUG_LOG'].print(f'[SEQ ERROR] {err_msg}')
            sg.popup_error(err_msg, keep_on_top=True)
        continue
    if event == 'ALL_STEP_TIME':
        payload = values.get(event, None)
        if isinstance(payload, (list, tuple)) and len(payload) >= 2:
            step_idx, elapsed = payload[0], payload[1]
            key = f'ALL_STEP{step_idx}_TIME'
            if key in window.AllKeysDict:
                try:
                    window[key].update(f"{float(elapsed):.2f}")
                except Exception:
                    window[key].update('—')
        continue
    if event == 'ALL_SEQ_DONE':
        status = values.get(event, '')
        SEQ_RUNNING = False
        SEQ_STOP_EVENT = None
        SEQ_THREAD = None
        if 'ALL_RUN_SEQUENCE' in window.AllKeysDict:
            window['ALL_RUN_SEQUENCE'].update(disabled=False)
        if 'ALL_STOP_SEQUENCE' in window.AllKeysDict:
            window['ALL_STOP_SEQUENCE'].update(disabled=True)
        if not window_closed:
            window['DEBUG_LOG'].print(f'[SEQ] Done: {status}')
            if isinstance(status, str):
                if status.startswith('error'):
                    sg.popup_error(status, keep_on_top=True)
                elif status == 'completed':
                    sg.popup_ok('Sequence complete.', keep_on_top=True)
        continue
    if event == 'ALL_STOP_SEQUENCE':
        if SEQ_STOP_EVENT is not None:
            SEQ_STOP_EVENT.set()
        continue

    # DataPipe events
    if event == 'DP_LOAD':
        file_path = values.get('DP_FILE', '')
        sheet_name = values.get('DP_SHEET', '') or None
        try:
            row_start = int(str(values.get('DP_ROW_START', '2')).strip() or '2')
            row_end = int(str(values.get('DP_ROW_END', '61')).strip() or '61')
        except Exception:
            row_start, row_end = 2, 61
        try:
            raw_segments, seconds_guess, missing_axes = load_datapipe_segments(file_path, sheet_name, row_start, row_end)
            prepared_segments = prepare_datapipe_segments(raw_segments)
            window._dp_segments = prepared_segments
            window._dp_time_ms = prepared_segments[0]['time_ms'] if prepared_segments else None
            render_datapipe_preview(window, prepared_segments)
            time_note = 'seconds converted to ms' if seconds_guess else 'ms'
            missing_note = f"; missing headers treated as 0: {', '.join(missing_axes)}" if missing_axes else ''
            window['DP_STATUS'].update(f"Loaded {len(prepared_segments)} segments ({time_note}{missing_note}).")
            if not window_closed and 'DEBUG_LOG' in window.AllKeysDict:
                window['DEBUG_LOG'].print(f"[DP_LOAD] missing axes: {missing_axes}")
            enable_dp = bool(prepared_segments)
            window['DP_SEND'].update(disabled=not enable_dp)
            if 'DP_SEND_PR' in window.AllKeysDict:
                window['DP_SEND_PR'].update(disabled=not enable_dp)
            if 'DP_SEND_BATCH_PR' in window.AllKeysDict:
                window['DP_SEND_BATCH_PR'].update(disabled=not enable_dp)
        except Exception as e:
            error_details = f"[DP_LOAD] {e}\n" + traceback.format_exc()
            window['DP_STATUS'].update(f"Load failed: {e}")
            if not window_closed and 'DEBUG_LOG' in window.AllKeysDict:
                window['DEBUG_LOG'].print(error_details)

    if event == 'DP_SEND_BATCH_PR':
        try:
            segments = getattr(window, '_dp_segments', None)
            if not segments:
                window['DEBUG_LOG'].print('No segments loaded for batch PR.')
            else:
                program_str = send_batch_pr_program(comm, segments)
                window['DEBUG_LOG'].print('Batch PR program sent and executed.')
        except Exception as e:
            window['DEBUG_LOG'].print(f'Error: {e}')
            window['DP_PREVIEW'].update('')
            window['DP_SEND'].update(disabled=True)
            if 'DP_SEND_PR' in window.AllKeysDict:
                window['DP_SEND_PR'].update(disabled=True)
        continue

    if event == 'DP_SEND':
        segments = getattr(window, '_dp_segments', None)
        try:
            if not segments:
                raise RuntimeError('No segments loaded. Load first.')
            send_datapipe_contour(comm, segments, window)
            window['DP_STATUS'].update('Contour data sent to controller (DT uses first segment).')
        except Exception as e:
            window['DP_STATUS'].update(f"Send failed: {e}")
        continue

    if event == 'DP_SEND_PR':
        segments = getattr(window, '_dp_segments', None)
        if not segments:
            window['DP_STATUS'].update('No segments loaded. Load first.')
            continue
        try:
            line_speed = float(str(values.get('ALL_LINE_SPEED', '1')).strip() or '1')
        except Exception:
            line_speed = 1.0
        if line_speed < 0:
            line_speed = 0.0
        max_rows = None
        try:
            max_rows_val = str(values.get('DP_RUN_ROWS', '')).strip()
            if max_rows_val:
                max_rows = int(float(max_rows_val))
                if max_rows <= 0:
                    max_rows = None
        except Exception:
            max_rows = None
        if not window_closed and 'DEBUG_LOG' in window.AllKeysDict:
            window['DEBUG_LOG'].print(f"[DP_SEND_PR] rows={len(segments)} max_rows={max_rows} line_speed={line_speed}")

        import threading
        DP_PR_STOP_EVENT = threading.Event()
        window._dp_pr_stop = DP_PR_STOP_EVENT
        window['DP_STATUS'].update('Sending PR sequence...')

        def _run_dp_pr():
            try:
                send_datapipe_pr(comm, segments, window, line_speed=line_speed, values=values, max_rows=max_rows, stop_event=DP_PR_STOP_EVENT)
                ran_rows = max_rows if (max_rows is not None and max_rows > 0) else len(segments)
                window.write_event_value('DP_PR_DONE', f'PR sequence sent (rows={ran_rows}).')
            except Exception as e:
                window.write_event_value('DP_PR_ERROR', f"Send failed: {e}")

        threading.Thread(target=_run_dp_pr, daemon=True).start()
        continue

    if event == 'DP_PR_PROGRESS':
        payload = values.get(event, None)
        if isinstance(payload, (list, tuple)) and len(payload) == 2 and 'DP_STATUS' in window.AllKeysDict:
            idx, total = payload
            window['DP_STATUS'].update(f'Sending PR: {idx}/{total}')
        continue

    if event == 'DP_PR_DONE':
        msg = values.get(event, '')
        if 'DP_STATUS' in window.AllKeysDict:
            window['DP_STATUS'].update(msg)
        continue

    if event == 'DP_PR_ERROR':
        msg = values.get(event, '')
        if 'DP_STATUS' in window.AllKeysDict:
            window['DP_STATUS'].update(msg)
        if not window_closed and 'DEBUG_LOG' in window.AllKeysDict:
            window['DEBUG_LOG'].print(f"[DP_SEND_PR] error: {msg}")
        continue

    if event == 'ALL_PVT_SEND':
        try:
            sample_ms = float(str(values.get('ALL_PVT_SAMPLE_MS', '50')).strip() or '50')
            if sample_ms <= 0:
                raise ValueError('Sample time must be positive.')
            payload = build_all_pvt_payload(values, window, sample_ms)
            window._pvt_payload = payload
            if 'PVT_PREVIEW' in window.AllKeysDict:
                render_pvt_preview(window, payload)
            send_pvt_payload(comm, payload, window)
            status_msg = f"Sent {payload['count']} PVT points from ALL tab @ {sample_ms:.1f} ms."
            if 'ALL_PVT_STATUS' in window.AllKeysDict:
                window['ALL_PVT_STATUS'].update(status_msg)
            if 'PVT_STATUS' in window.AllKeysDict:
                window['PVT_STATUS'].update(status_msg)
            if not window_closed and 'DEBUG_LOG' in window.AllKeysDict:
                window['DEBUG_LOG'].print(f"[ALL_PVT_SEND] points={payload['count']} sample={sample_ms}")
        except Exception as e:
            fail_msg = f"Send failed: {e}"
            if 'ALL_PVT_STATUS' in window.AllKeysDict:
                window['ALL_PVT_STATUS'].update(fail_msg)
            if 'PVT_STATUS' in window.AllKeysDict:
                window['PVT_STATUS'].update(fail_msg)
            if not window_closed and 'DEBUG_LOG' in window.AllKeysDict:
                window['DEBUG_LOG'].print(f"[ALL_PVT_SEND] error: {e}\n{traceback.format_exc()}")
        continue

    if event == 'PVT_LOAD':
        file_path = str(values.get('PVT_FILE', '') or '').strip()
        try:
            sample_ms = float(str(values.get('PVT_SAMPLE_MS', '50')).strip() or '50')
        except Exception:
            sample_ms = 50.0
        try:
            if not file_path:
                raise ValueError('Select a PVT file first.')
            if sample_ms <= 0:
                raise ValueError('Sample time must be positive.')
            raw_rows = load_pvt_points(file_path)
            payload = prepare_pvt_payload(raw_rows, sample_ms)
            window._pvt_payload = payload
            render_pvt_preview(window, payload)
            window['PVT_STATUS'].update(f"Loaded {payload['count']} points @ {sample_ms:.1f} ms.")
            window['PVT_SEND'].update(disabled=False)
            if not window_closed and 'DEBUG_LOG' in window.AllKeysDict:
                window['DEBUG_LOG'].print(f"[PVT_LOAD] points={payload['count']} sample={sample_ms} file={file_path}")
        except Exception as e:
            window['PVT_STATUS'].update(f"Load failed: {e}")
            window['PVT_PREVIEW'].update('')
            if 'PVT_SEND' in window.AllKeysDict:
                window['PVT_SEND'].update(disabled=True)
            if not window_closed and 'DEBUG_LOG' in window.AllKeysDict:
                window['DEBUG_LOG'].print(f"[PVT_LOAD] error: {e}\n{traceback.format_exc()}")
        continue

    if event == 'PVT_SEND':
        payload = getattr(window, '_pvt_payload', None)
        try:
            if not payload:
                raise RuntimeError('Load PVT data first.')
            send_pvt_payload(comm, payload, window)
            window['PVT_STATUS'].update('PVT sent to controller (PT/PV/PA, axes A-D).')
        except Exception as e:
            window['PVT_STATUS'].update(f"Send failed: {e}")
            if not window_closed and 'DEBUG_LOG' in window.AllKeysDict:
                window['DEBUG_LOG'].print(f"[PVT_SEND] error: {e}\n{traceback.format_exc()}")
        continue
    # Ensure counters are initialized before use
    if not hasattr(window, '_invalid_resp_counters'):
        window._invalid_resp_counters = [0]*8
    if not hasattr(window, '_consecutive_zero_actuals'):
        window._consecutive_zero_actuals = [0]*8
    if not hasattr(window, '_last_valid_pos'):
        window._last_valid_pos = ['']*8
    if not hasattr(window, '_last_pos_update_ts'):
        window._last_pos_update_ts = [None]*8
    if not hasattr(window, '_limit_tripped'):
        window._limit_tripped = [False]*8
    if not hasattr(window, '_limit_exceed_counts'):
        window._limit_exceed_counts = [0]*8
    if not hasattr(window, '_jog_limit_hit'):
        window._jog_limit_hit = [False]*8

    # Sync per-servo description input to ALL tab label
    if isinstance(event, str) and event.startswith('S') and event.endswith('_desc'):
        try:
            servo_num = int(event[1:event.index('_')])
            desc_text = str(values.get(event, '')).strip()
            display_text = desc_text if desc_text else DEFAULT_SERVO_DESCRIPTIONS.get(servo_num, f'Servo {servo_num}')
            if f'ALL_S{servo_num}_desc' in window.AllKeysDict:
                window[f'ALL_S{servo_num}_desc'].update(display_text)
            axis_letter = AXIS_LETTERS[servo_num - 1]
            AXIS_UNITS.setdefault(axis_letter, {})['description'] = display_text
            save_axis_description(axis_letter, display_text)
        except Exception as desc_err:
            if not window_closed:
                window['DEBUG_LOG'].print(f'[ERROR] Descriptor update failed: {desc_err}\n{traceback.format_exc()}')
        continue

    if handle_all_tab_event(window, event, values):
        continue
    if event == 'ALL_RUN_SEQUENCE':
        handle_all_run_sequence(window, comm, values)
        continue

    if event == 'JOG_PRESS':
        try:
            servo_num, direction, is_press = values.get(event, (None, None, None))
        except Exception:
            servo_num, direction, is_press = None, None, None
        if servo_num is not None and direction is not None and is_press is not None:
            handle_jog_press(window, int(servo_num), direction, bool(is_press), values)
        else:
            print(f'[DEBUG] Invalid JOG_PRESS payload: {values.get(event)}')
        continue

    if event == 'JOG_LIMIT_HIT':
        try:
            servo_num, which_limit = values.get(event, (None, None))
            if servo_num is not None:
                if hasattr(window, '_jog_limit_hit'):
                    window._jog_limit_hit[int(servo_num) - 1] = True
                if str(which_limit).lower() == 'max':
                    window[f'S{servo_num}_status_light'].update('●', text_color='#FFA500')
                    window[f'S{servo_num}_status_text'].update('At Max Limit', text_color='#FFA500')
                else:
                    window[f'S{servo_num}_status_light'].update('●', text_color='#FFA500')
                    window[f'S{servo_num}_status_text'].update('At Min Limit', text_color='#FFA500')
        except Exception as jog_limit_err:
            print(f'[DEBUG] Failed to handle JOG_LIMIT_HIT: {jog_limit_err}')
        continue


    if event == sg.WIN_CLOSED:
        window_closed = True
        if SEQ_STOP_EVENT is not None:
            try:
                SEQ_STOP_EVENT.set()
            except Exception:
                pass
        try:
            save_sequence_state_from_values(values)
        except Exception:
            pass
        try:
            # [CHANGE 2026-03-24 11:19:00 -04:00] Persist speed/accel/decel for next startup.
            save_motion_defaults_from_values(values)
        except Exception:
            pass
        break

    # Logging errors/warnings forwarded from background threads and communications.py
    if event == 'GUI_LOG':
        msg = values.get('GUI_LOG', '')
        if msg and not window_closed:
            window['DEBUG_LOG'].print(msg)

    # [CHANGE 2026-04-17 00:00:00 -04:00] Comm health: update per-axis link indicator dots.
    if event == 'COMM_HEALTH':
        health_data = values.get('COMM_HEALTH', {})
        if isinstance(health_data, dict) and not window_closed:
            for servo_num, info in health_data.items():
                ok = info.get('ok')
                label = info.get('label', '')
                ind_key   = f'S{servo_num}_comm_indicator'
                label_key = f'S{servo_num}_comm_label'
                if ind_key not in window.AllKeysDict:
                    continue
                if ok is True:
                    color = '#00CC00'   # green
                elif ok is False:
                    color = '#FF3333'   # red
                else:
                    color = 'gray'      # not configured
                window[ind_key].update('\u25cf', text_color=color)
                window[label_key].update(label, text_color=color)
        continue

    if event == 'POSITION_POLL':
        # Handle position update from background thread
        data = values[event] if event in values else None
        if data:
            i = data['servo']
            axis_letter = data['axis_letter']
            pos_resp = data.get('pos_resp')
            raw_resp = data.get('raw_resp')
            pos_val = None
            pos_val_deg = None
            valid = False
            # [CHANGE 2026-03-27 11:35:00 -04:00] Axis E has no encoder; render Actual from commanded cache only.
            axis_e_allow_update = True
            axis_e_override_pulses = None
            if axis_letter == 'E':
                try:
                    if comm_e is not None:
                        commanded = getattr(comm_e, 'clearcore_commanded_position', None)
                        if commanded is None:
                            commanded = getattr(comm_e, 'clearcore_last_position', None)
                        if commanded is not None:
                            axis_e_override_pulses = float(commanded)
                            pos_resp = str(commanded)
                except Exception:
                    pass
            # Log the raw response for debugging (optional)
            if not window_closed and LOG_POSITION_POLLS:
                window['DEBUG_LOG'].print(f'Axis {axis_letter}: MG _RP{axis_letter} raw response: {raw_resp}')
            # [CHANGE 2026-03-24 15:40:00 -04:00] Disabled per-poll Servo 5 button-state query to prevent comm congestion/delays.
            if axis_letter == 'E' and axis_e_allow_update and (pos_resp is None or str(pos_resp).strip() in (':', '')):
                try:
                    if comm_e is not None:
                        commanded = getattr(comm_e, 'clearcore_commanded_position', None)
                        if commanded is not None:
                            pos_resp = str(commanded)
                except Exception:
                    pass
            if axis_e_override_pulses is not None:
                pos_resp = str(axis_e_override_pulses)
            if axis_e_allow_update and pos_resp is not None and str(pos_resp).strip() != ':' and str(pos_resp).strip() != '':
                try:
                    pos_val_pulses = float(pos_resp)
                    axis_units = AXIS_UNITS[axis_letter]
                    pulses_per_degree = axis_units.get('scaling') or (axis_units.get('pulses', 0) / max(axis_units.get('degrees', 1), 1e-9))
                    if pulses_per_degree <= 0:
                        pulses_per_degree = 1
                    gearbox = axis_units.get('gearbox', 1)
                    pos_val_deg = pos_val_pulses / (pulses_per_degree * gearbox)
                    pos_val_disp = 0 if abs(pos_val_deg) < 1e-6 else round(pos_val_deg, 2)
                    # Robust: Only treat zero as valid if setpoint was zero or after 3 consecutive zero responses
                    last_setpoint = getattr(window, '_last_setpoints', [{}]*8)[i-1].get('abs_pos', None)
                    try:
                        last_setpoint_zero = last_setpoint is not None and abs(float(last_setpoint)) < 1e-6
                    except Exception:
                        last_setpoint_zero = False
                    consecutive_zero = getattr(window, '_consecutive_zero_actuals', [0]*8)
                    # Suppress a zero reading only if we've previously seen a non-zero position
                    # (motor was somewhere non-zero and is now falsely reading 0 during decel/stop).
                    # Before the first real move, 0 is genuine and must be shown.
                    # After 3 consecutive zeros, accept as genuine (e.g. motor parked at 0° end-stop).
                    last_known = window._last_valid_pos[i-1] if window._last_valid_pos[i-1] else ''
                    has_seen_nonzero = last_known not in ('', '0', 'N/A')
                    if pos_val_disp == 0 and not last_setpoint_zero and has_seen_nonzero and consecutive_zero[i-1] < 3:
                        valid = False
                    else:
                        if not window_closed:
                            window[f'S{i}_actual_pos'].update(str(pos_val_disp))
                            update_setpoint_highlight(window, i, pos_val_deg)
                        window._last_valid_pos[i-1] = str(pos_val_disp)
                        window._last_pos_update_ts[i-1] = time.time()
                        window._invalid_resp_counters[i-1] = 0
                        valid = True
                except Exception:
                    pass
            # Actual position in pulses display
            if pos_resp is not None and str(pos_resp).strip() not in (':', ''):
                try:
                    pos_pulses_disp = str(pos_resp).strip()
                    if not window_closed:
                        window[f'S{i}_actual_pos_pulses'].update(pos_pulses_disp)
                except Exception:
                    if not window_closed:
                        window[f'S{i}_actual_pos_pulses'].update('N/A')
            # SAFETY: Stop motion if position exceeds soft limits OR absolute 360-degree rotation limit
            if pos_val_deg is not None and not SAFETY_LIMIT_STOPS_ENABLED:
                window._limit_exceed_counts[i-1] = 0
                window._limit_tripped[i-1] = False
                window._jog_limit_hit[i-1] = False
            if pos_val_deg is not None and SAFETY_LIMIT_STOPS_ENABLED:
                axis_units = AXIS_UNITS[axis_letter]
                min_val = axis_units['min']
                max_val = axis_units['max']
                # Absolute safety: never allow >360 degrees rotation
                beyond_absolute_limit = abs(pos_val_deg) > ABSOLUTE_SAFETY_LIMIT_DEG
                beyond_soft_limit = (pos_val_deg < min_val - LIMIT_SOFT_TOLERANCE_DEG or pos_val_deg > max_val + LIMIT_SOFT_TOLERANCE_DEG)
                
                if beyond_soft_limit or beyond_absolute_limit:
                    window._limit_exceed_counts[i-1] += 1
                    if not window_closed:
                        window['DEBUG_LOG'].print(f'[LIMIT] Axis {axis_letter} out-of-range sample {window._limit_exceed_counts[i-1]}/{LIMIT_TRIP_CONFIRM_SAMPLES}: pos={pos_val_deg:.3f}° (soft {min_val-LIMIT_SOFT_TOLERANCE_DEG:.1f}..{max_val+LIMIT_SOFT_TOLERANCE_DEG:.1f}, abs±{ABSOLUTE_SAFETY_LIMIT_DEG:.0f})')
                else:
                    window._limit_exceed_counts[i-1] = 0

                if (beyond_soft_limit or beyond_absolute_limit) and not window._limit_tripped[i-1] and window._limit_exceed_counts[i-1] >= LIMIT_TRIP_CONFIRM_SAMPLES:
                    stop_key = f'S{i}_stop'
                    stop_cmd = COMMAND_MAP.get(stop_key)
                    controller = get_comm_for_axis(axis_letter)
                    # [CHANGE 2026-03-24 16:24:00 -04:00] Route safety limit-stop through per-axis comm path so E/H stop on their native controllers.
                    if stop_cmd and controller:
                        try:
                            stop_cmd_val = stop_cmd if not callable(stop_cmd) else stop_cmd()
                            send_axis_command(axis_letter, stop_cmd_val)
                            # Clear motion command tracking
                            LAST_MOTION_COMMAND[i-1] = None
                            if not window_closed:
                                if beyond_absolute_limit:
                                    limit_msg = f'[SAFETY] Axis {axis_letter} exceeded ABSOLUTE 360° rotation limit at {pos_val_deg:.1f}°; EMERGENCY STOP sent: {stop_cmd_val}'
                                    popup_msg = f'EMERGENCY STOP!\n\nAxis {axis_letter} exceeded absolute safety limit.\nPosition: {pos_val_deg:.1f}°\n\nServos must NEVER rotate more than 360°.'
                                else:
                                    limit_msg = f'[WARN] Axis {axis_letter} exceeded soft limits ({min_val},{max_val}); sent stop command: {stop_cmd_val}'
                                    popup_msg = f'Axis {axis_letter} exceeded limits ({min_val} to {max_val}). Motion stopped.'
                                window['DEBUG_LOG'].print(limit_msg)
                                # Visual + popup notification on first limit trip
                                window[f'S{i}_status_light'].update('●', text_color='#FF4500')  # Orange-red
                                window[f'S{i}_status_text'].update('Stopped (limit)', text_color='#FF4500')
                                sg.popup_ok(popup_msg, keep_on_top=True, title='Safety Stop' if beyond_absolute_limit else '')
                        except Exception:
                            if not window_closed:
                                window['DEBUG_LOG'].print(f'[ERROR] Failed to send stop for axis {axis_letter}')
                    elif not controller and not window_closed:
                        window['DEBUG_LOG'].print(f'[WARN] Axis {axis_letter} exceeded limits but comm not initialized; no stop sent')
                        sg.popup_ok(f'Axis {axis_letter} exceeded limits ({min_val} to {max_val}) but comm not initialized; stop not sent.', keep_on_top=True, title='')
                    window._limit_tripped[i-1] = True
                elif window._limit_tripped[i-1] and min_val <= pos_val_deg <= max_val:
                    # Clear limit indicator when back inside bounds
                    window._limit_tripped[i-1] = False
                    window._limit_exceed_counts[i-1] = 0
                    if not window_closed:
                        window[f'S{i}_status_light'].update('●', text_color='#00FF00')
                        window[f'S{i}_status_text'].update('Enabled', text_color='#00FF00')
                elif window._jog_limit_hit[i-1] and min_val < pos_val_deg < max_val:
                    # Clear jog limit indicator when back inside absolute bounds
                    window._jog_limit_hit[i-1] = False
                    window._limit_exceed_counts[i-1] = 0
                    if not window_closed:
                        window[f'S{i}_status_light'].update('●', text_color='#00FF00')
                        window[f'S{i}_status_text'].update('Enabled', text_color='#00FF00')
            # Robust actuals display: only accept zero if setpoint was zero or after 3 consecutive zero responses
            debug_msgs = []
            if not valid:
                window._invalid_resp_counters[i-1] += 1
                last_setpoint = getattr(window, '_last_setpoints', [{}]*8)[i-1].get('abs_pos', None)
                try:
                    last_setpoint_zero = last_setpoint is not None and abs(float(last_setpoint)) < 1e-6
                except Exception:
                    last_setpoint_zero = False
                consecutive_zero = getattr(window, '_consecutive_zero_actuals', [0]*8)
                if pos_val_disp == 0:
                    consecutive_zero[i-1] = consecutive_zero[i-1] + 1
                else:
                    consecutive_zero[i-1] = 0
                window._consecutive_zero_actuals = consecutive_zero
                debug_msgs.append(f'[DEBUG] Axis {axis_letter} raw={pos_resp} disp={pos_val_disp} setpoint={last_setpoint} zero_ctr={consecutive_zero[i-1]} valid={valid}')
                if window._invalid_resp_counters[i-1] >= 5:
                    if not window_closed:
                        window[f'S{i}_actual_pos'].update('N/A')
                    log_val = 'N/A'
                    debug_msgs.append(f'[DEBUG] Axis {axis_letter} display updated to N/A (invalid_ctr={window._invalid_resp_counters[i-1]})')
                else:
                    last_val = window._last_valid_pos[i-1] if window._last_valid_pos[i-1] else ''
                    if not window_closed:
                        window[f'S{i}_actual_pos'].update(last_val)
                    log_val = last_val if last_val else 'N/A'
                    debug_msgs.append(f'[DEBUG] Axis {axis_letter} display kept at last valid ({last_val})')
            else:
                if hasattr(window, '_consecutive_zero_actuals'):
                    window._consecutive_zero_actuals[i-1] = 0
                log_val = window._last_valid_pos[i-1]
                debug_msgs.append(f'[DEBUG] Axis {axis_letter} valid actual: {log_val}')
            if not window_closed and LOG_POSITION_POLLS:
                for msg in debug_msgs:
                    window['DEBUG_LOG'].print(msg)

            if not window_closed and LOG_POSITION_POLLS:
                window['DEBUG_LOG'].print(f'Axis {axis_letter}: MG _RP{axis_letter} -> {log_val}')
        continue
    # Reconnect button — re-establishes the TCP/UDP link for this axis's controller
    if isinstance(event, str) and event.endswith('_reconnect'):
        try:
            servo_num = int(event[1:event.index('_')])
            axis_letter = AXIS_LETTERS[servo_num - 1]
            target_comm = get_comm_for_axis(axis_letter)
            if target_comm is not None and hasattr(target_comm, '_reconnect_rsi'):
                success = target_comm._reconnect_rsi()
                msg = 'Reconnected successfully.' if success else 'Reconnect failed — check service is running.'
            elif target_comm is not None:
                msg = 'Link uses UDP (ClearCore) — no explicit reconnect needed.'
            else:
                msg = 'No comm object configured for this axis.'
            window['DEBUG_LOG'].print(f'[RECONNECT] Axis {axis_letter}: {msg}')
        except Exception as _re:
            window['DEBUG_LOG'].print(f'[RECONNECT] Error: {_re}')

    # Zero Position button (handle early so it isn't swallowed by generic S*_action logic)
    if isinstance(event, str) and event.endswith('_zero_pos'):
        try:
            servo_num = int(event[1:event.index('_')])
            axis_letter = AXIS_LETTERS[servo_num - 1]
            confirm = sg.popup_yes_no(
                f'Set current position as ZERO for Axis {axis_letter}?\n\nThis cannot be undone without re-homing.',
                title='Confirm Zero Position',
                keep_on_top=True,
            )
            if confirm != 'Yes':
                continue

            # Use axis-specific command instead of multi-axis format
            dp_cmd = f"DP{axis_letter}=0"

            # Route to correct comm object
            response = send_axis_command(axis_letter, dp_cmd)
            if axis_letter == 'E':
                # [CHANGE 2026-03-27 11:20:00 -04:00] Servo E zero immediately resets displayed/cached commanded position.
                zero_ok = not (
                    response is False or
                    (isinstance(response, str) and str(response).strip().upper().startswith('UNSUPPORTED'))
                )
                if zero_ok:
                    try:
                        if comm_e is not None:
                            setattr(comm_e, 'clearcore_last_position', 0)
                            setattr(comm_e, 'clearcore_commanded_position', 0)
                        sync_axis_e_actual_from_commanded(window, servo_num)
                    except Exception:
                        pass
        except Exception as e:
            sg.popup_error(f'Error parsing servo number: {e}', keep_on_top=True)
        continue

    # Keypad button logic for numeric value entry
    if event in NUMERIC_KEYPAD_BUTTONS:
        parts = event.split('_')
        servo = parts[0]
        field = '_'.join(parts[1:-1])
        input_key = f'{servo}_{field}'
        current_val = values.get(input_key, '')
        try:
            current_val = round(float(current_val), 1)
        except (ValueError, TypeError):
            current_val = 0.0
        # Convert servo number to axis letter (1->A, 2->B, ...)
        axis_num = int(servo[1:]) if servo.startswith('S') else None
        axis_letter = chr(64 + axis_num) if axis_num and 1 <= axis_num <= 8 else None
        # Use axis-specific limits (speed/accel/decel/positions) with defaults
        if axis_letter:
            min_val, max_val = get_limits(axis_letter, field)
        else:
            min_val, max_val = NUMERIC_LIMITS.get(field, (0, 54000))
        unit_label = 'DPS' if field == 'speed' else ('DPS^2' if field in ['accel', 'decel'] else ('Deg' if field in ['abs_pos', 'rel_pos', 'jog_amount'] else ''))
        # Make popup title more descriptive with setpoint type
        field_titles = {
            'speed': 'Speed',
            'accel': 'Acceleration',
            'decel': 'Deceleration',
            'abs_pos': 'Absolute Position',
            'rel_pos': 'Relative Position',
            'jog_amount': 'Jog Amount',
        }
        field_title = field_titles.get(field, field.capitalize())
        popup_title = f'Enter {field_title} for Servo {servo} ({unit_label})'
        keypad = NumericKeypad(
            title=popup_title,
            current_value=current_val,
            axis_letter=axis_letter,
            font=GLOBAL_FONT,
            unit_label=unit_label,
            min_val=min_val,
            max_val=max_val
        )
        result = keypad.show()
        if result is not None:
            # Enforce min/max for PC keyboard edits as well
            if result < min_val or result > max_val:
                sg.popup_error(f"Value for {field} must be between {min_val} and {max_val}", keep_on_top=True)
            else:
                window[input_key].update(format_display_value(result))
                # Mark pending (not confirmed) until OK is pressed
                try:
                    serv_num = int(servo[1:])
                    if field == 'abs_pos':
                        update_setpoint_highlight(window, serv_num)
                    else:
                        set_pending_highlight(window, serv_num, field)
                    if field in ('speed', 'accel', 'decel'):
                        update_mid_speed_display(window, serv_num)
                except Exception:
                    pass
        continue

    # Restrict keyboard entries for numeric fields to digits, leading '-', and a single decimal (one digit precision)
    if event in NUMERIC_INPUT_KEYS:
        val = values.get(event, '')
        import re
        # Keep only digits, '-', and '.'
        filtered = re.sub(r'[^0-9\-.]', '', val)
        # Normalize sign to leading position only
        sign = '-' if filtered.startswith('-') else ''
        filtered = filtered[1:] if filtered.startswith('-') else filtered
        filtered = filtered.replace('-', '')
        # Enforce a single decimal point and only one digit after it
        if '.' in filtered:
            whole, frac = filtered.split('.', 1)
            frac = frac[:1]
            filtered = f"{sign}{whole}.{frac}"
        else:
            filtered = sign + filtered
        # If the filtered value differs from the input, update the field
        if filtered != val:
            window[event].update(filtered)
        # Clamp to axis min/max for PC keyboard entry (same limits as keypad)
        if filtered not in ('', '-', '.', '-.'):
            try:
                numeric_val = float(filtered)
                parts = event.split('_', 1)
                servo_part = parts[0] if len(parts) > 0 else ''
                field_part = parts[1] if len(parts) > 1 else ''
                servo_num = None
                if servo_part.startswith('S'):
                    servo_num = int(servo_part[1:])
                elif servo_part == 'ALL' and field_part:
                    nested = field_part.split('_', 1)
                    nested_servo = nested[0] if nested else ''
                    if nested_servo.startswith('S'):
                        try:
                            servo_num = int(nested_servo[1:])
                        except Exception:
                            servo_num = None
                    field_part = 'abs_pos'
                axis_letter = AXIS_LETTERS[servo_num - 1] if servo_num and 1 <= servo_num <= 8 else None
                if axis_letter:
                    min_val, max_val = get_limits(axis_letter, field_part)
                else:
                    min_val, max_val = NUMERIC_LIMITS.get(field_part, (0, 54000))
                clamped = round(max(min_val, min(max_val, numeric_val)), 1)
                if clamped != numeric_val:
                    window[event].update(format_display_value(clamped))
                if servo_num and field_part:
                    if field_part == 'abs_pos':
                        update_setpoint_highlight(window, servo_num)
                    else:
                        set_pending_highlight(window, servo_num, field_part)
                    if field_part in ('speed', 'accel', 'decel'):
                        update_mid_speed_display(window, servo_num)
            except Exception:
                pass
    # Only call handle_servo_event for setpoint OK buttons
    if isinstance(event, str) and event.startswith('S') and '_' in event:
        if event.endswith('_ok'):
            print(f'[DEBUG] Main loop routing event to handle_servo_event: {event}')
            handle_servo_event(event, values)
            # Add polling pause after setpoint changes
            import time
            time.sleep(1)  # Pause polling for 1 second after setpoint change
            continue
        # Otherwise, handle direct motor control buttons (Enable, Disable, Start, Stop, Jog) by reusing handle_servo_event
        # to ensure a single code path with consistent scaling logic.
        parts = event.split('_')
        if len(parts) == 3 and parts[1] == 'jog' and parts[2] in ('cw', 'ccw'):
            # [CHANGE 2026-03-24 13:36:00 -04:00] Safety: jog buttons are one-shot pulses only (no release event dependency).
            servo_num = parts[0][1:]
            if not str(servo_num).isdigit():
                continue
            servo_num_int = int(servo_num)
            direction = parts[2]
            axis_letter = AXIS_LETTERS[servo_num_int - 1]
            window['DEBUG_LOG'].print(f'Button clicked: S{servo_num}_jog_{direction} (Axis {axis_letter}) [one-shot]')
            handle_jog_press(window, servo_num_int, direction, True, values)
            continue
        if len(parts) == 2:
            servo_num = parts[0][1:]
            if not str(servo_num).isdigit():
                continue
            action = parts[1]
            axis_letter = AXIS_LETTERS[int(servo_num)-1]
            prev_log = window['DEBUG_LOG'].get()
            if not prev_log.endswith('\r\n') and not prev_log.endswith('\n'):
                prev_log += '\r\n'
            window['DEBUG_LOG'].print(f'Button clicked: S{servo_num}_{action} (Axis {axis_letter})')
            # Reuse the unified handler (handles jog scaling to pulses)
            handle_servo_event(event, values)
        continue
